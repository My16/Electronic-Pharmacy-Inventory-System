# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import (
    Q, Sum, Count, F, Value, DecimalField,
    ExpressionWrapper, FloatField, Prefetch
)
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import (
    UserProfile, UserPermission, Medicine, MedicineCategory, MedicineStock,
    Supplier, StockMovement, Dispensing, DispensingItem, Doctor
)
import json
import io

from django.views.decorators.http   import require_GET

from .models          import Dispensing, DispensingItem
from .ihomis_models   import HPerson

from datetime import date, datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


from django.db import transaction
import calendar

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_role(user):
    if user.is_superuser:
        return 'admin'
    try:
        return user.profile.role or 'user'
    except AttributeError:
        return 'user'

def is_admin(user):
    return user.is_superuser or get_role(user) == 'admin'

def can_access_module(user, module):
    """
    Admins/superusers always pass.
    Everyone else checks their UserPermission row.
    """
    if user.is_superuser or get_role(user) == 'admin':
        return True
    try:
        return getattr(user.permissions, module, False)
    except Exception:
        return False


def _deduct_fefo(medicine, qty_needed, performed_by,
                 movement_type='out', dispensed_to='',
                 prescription_no='', notes='', reference_no=''):
    """
    Deduct `qty_needed` units from `medicine` using FEFO
    (earliest expiry first, then batches with no expiry).
    Creates one StockMovement log entry per batch touched.
    Returns True on success, raises ValueError if insufficient stock.
    """
    total_available = medicine.stock_quantity
    if qty_needed > total_available:
        raise ValueError(
            f'Insufficient stock for "{medicine.medicine_name}". '
            f'Available: {total_available}.'
        )

    # Batches ordered: soonest expiry first, then no-expiry batches, then oldest received
    batches = list(
        medicine.batches.filter(quantity__gt=0)
        .order_by(
            F('expiry_date').asc(nulls_last=True),
            'received_at'
        )
    )

    remaining = qty_needed
    for batch in batches:
        if remaining <= 0:
            break
        take = min(batch.quantity, remaining)
        old_total = medicine.stock_quantity          # before this batch deduction
        batch.quantity -= take
        batch.save()
        new_total = medicine.stock_quantity          # after

        StockMovement.objects.create(
            medicine        = medicine,
            batch           = batch,
            movement_type   = movement_type,
            quantity        = -take,
            quantity_before = old_total,
            quantity_after  = new_total,
            batch_number    = batch.batch_number,
            expiry_date     = batch.expiry_date,
            purchase_price  = batch.purchase_price,
            supplier        = batch.supplier,
            dispensed_to    = dispensed_to,
            prescription_no = prescription_no,
            reference_no    = reference_no,
            notes           = notes,
            performed_by    = performed_by,
        )
        remaining -= take

    return True


def _parse_date_range(request, default_days=30):
    today     = timezone.now().date()
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    try:
        date_from = date.fromisoformat(date_from)
    except (ValueError, TypeError):
        date_from = today - timedelta(days=default_days)
    try:
        date_to = date.fromisoformat(date_to)
    except (ValueError, TypeError):
        date_to = today
    return date_from, date_to


# ─────────────────────────────────────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        if not username or not password:
            messages.error(request, 'Please enter both username and password.')
            return render(request, 'loginpage.html')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                return redirect(request.GET.get('next', 'dashboard'))
            messages.error(request, 'Your account has been disabled.')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'loginpage.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def dashboard_view(request):
    from dateutil.relativedelta import relativedelta

    today            = timezone.now().date()
    expiry_threshold = today + timezone.timedelta(days=90)

    active_medicines = Medicine.objects.filter(is_active=True)
    total_medicines  = active_medicines.count()

    # Low / out counts use annotated total stock per medicine
    med_with_stock = active_medicines.annotate(
        total_stock=Sum('batches__quantity')
    )
    low_stock_count    = med_with_stock.filter(
                             total_stock__gt=0,
                             total_stock__lte=F('reorder_level')).count()
    out_of_stock_count = med_with_stock.filter(
                             total_stock__isnull=True).count() + \
                         med_with_stock.filter(total_stock=0).count()

    expiring_soon_count = MedicineStock.objects.filter(
                              medicine__is_active=True,
                              quantity__gt=0,
                              expiry_date__lte=expiry_threshold,
                              expiry_date__gte=today).values('medicine').distinct().count()

    # Total stock value: sum(selling_price * batch_quantity) per medicine
    stock_value_qs = MedicineStock.objects.filter(
        medicine__is_active=True, quantity__gt=0
    ).annotate(
        line_value=ExpressionWrapper(
            F('medicine__selling_price') * F('quantity'),
            output_field=DecimalField()
        )
    ).aggregate(total=Sum('line_value'))
    total_stock_value = stock_value_qs['total'] or 0

    today_qs              = StockMovement.objects.filter(created_at__date=today)
    total_in_today        = today_qs.filter(movement_type__in=('in', 'return')).aggregate(
                                t=Sum('quantity'))['t'] or 0
    total_out_today       = abs(today_qs.filter(movement_type__in=('out', 'expired', 'damaged')).aggregate(
                                t=Sum('quantity'))['t'] or 0)
    total_movements_today = today_qs.count()

    recent_movements = StockMovement.objects.select_related(
                           'medicine', 'performed_by').order_by('-created_at')[:8]

    # Low stock items — medicines whose total batch qty <= reorder_level
    low_stock_items = []
    for m in active_medicines.prefetch_related('batches'):
        if 0 < m.stock_quantity <= m.reorder_level:
            low_stock_items.append(m)
        if len(low_stock_items) >= 8:
            break

    expiring_items = MedicineStock.objects.filter(
        medicine__is_active=True,
        quantity__gt=0,
        expiry_date__lte=expiry_threshold,
        expiry_date__gte=today
    ).select_related('medicine').order_by('expiry_date')[:6]

    monthly_labels, monthly_in, monthly_out = [], [], []
    for i in range(5, -1, -1):
        month_start = today.replace(day=1) - relativedelta(months=i)
        month_end   = month_start + relativedelta(months=1)
        qs          = StockMovement.objects.filter(
                          created_at__date__gte=month_start,
                          created_at__date__lt=month_end)
        in_qty  = qs.filter(movement_type__in=('in', 'return')).aggregate(
                      t=Sum('quantity'))['t'] or 0
        out_qty = abs(qs.filter(movement_type__in=('out', 'expired', 'damaged')).aggregate(
                      t=Sum('quantity'))['t'] or 0)
        monthly_labels.append(month_start.strftime('%b %Y'))
        monthly_in.append(in_qty)
        monthly_out.append(out_qty)

    weekly_labels, weekly_in, weekly_out = [], [], []
    for i in range(6, -1, -1):
        day     = today - timezone.timedelta(days=i)
        qs      = StockMovement.objects.filter(created_at__date=day)
        in_qty  = qs.filter(movement_type__in=('in', 'return')).aggregate(
                      t=Sum('quantity'))['t'] or 0
        out_qty = abs(qs.filter(movement_type__in=('out', 'expired', 'damaged')).aggregate(
                      t=Sum('quantity'))['t'] or 0)
        weekly_labels.append(day.strftime('%a %d'))
        weekly_in.append(in_qty)
        weekly_out.append(out_qty)

    context = {
        'total_medicines':       total_medicines,
        'low_stock_count':       low_stock_count,
        'out_of_stock_count':    out_of_stock_count,
        'expiring_soon_count':   expiring_soon_count,
        'total_stock_value':     total_stock_value,
        'total_in_today':        total_in_today,
        'total_out_today':       total_out_today,
        'total_movements_today': total_movements_today,
        'recent_movements':      recent_movements,
        'low_stock_items':       low_stock_items,
        'expiring_items':        expiring_items,
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_in':     json.dumps(monthly_in),
        'monthly_out':    json.dumps(monthly_out),
        'weekly_labels':  json.dumps(weekly_labels),
        'weekly_in':      json.dumps(weekly_in),
        'weekly_out':     json.dumps(weekly_out),
    }
    return render(request, 'dashboard.html', context)


# ─────────────────────────────────────────────────────────────────────────────
# INVENTORY
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'inventory'), login_url='dashboard')
def inventory_view(request):
    today            = timezone.now().date()
    expiry_threshold = today + timezone.timedelta(days=90)

    search        = request.GET.get('search', '').strip()
    category_id   = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    dosage_filter = request.GET.get('dosage', '')
    highlight_pk  = request.GET.get('highlight', '')
    per_page      = request.GET.get('per_page', '15')
    try:
        per_page = int(per_page)
        if per_page not in [10, 15, 25, 50, 100]:
            per_page = 15
    except ValueError:
        per_page = 15

    medicines_qs = Medicine.objects.select_related('category', 'supplier').filter(
        is_active=True
    ).prefetch_related('batches')

    if search:
        medicines_qs = medicines_qs.filter(
            Q(medicine_name__icontains=search) | Q(generic_name__icontains=search) |
            Q(brand_name__icontains=search)    | Q(barcode__icontains=search)       |
            Q(batches__batch_number__icontains=search)
        ).distinct()
    if category_id:
        medicines_qs = medicines_qs.filter(category_id=category_id)
    if dosage_filter:
        medicines_qs = medicines_qs.filter(dosage_form=dosage_filter)

    # Status filters require Python-level evaluation because stock_quantity is a property
    if status_filter in ('low_stock', 'out_of_stock', 'expiring', 'expired'):
        all_meds = list(medicines_qs)
        if status_filter == 'low_stock':
            medicines_qs = [m for m in all_meds if m.is_low_stock]
        elif status_filter == 'out_of_stock':
            medicines_qs = [m for m in all_meds if m.is_out_of_stock]
        elif status_filter == 'expiring':
            medicines_qs = [m for m in all_meds if m.is_expiring_soon and not m.is_expired]
        elif status_filter == 'expired':
            medicines_qs = [m for m in all_meds if m.is_expired]

    if highlight_pk:
        try:
            highlighted  = [m for m in medicines_qs if str(m.pk) == str(highlight_pk)]
            rest         = [m for m in medicines_qs if str(m.pk) != str(highlight_pk)]
            medicines_qs = highlighted + rest
        except Exception:
            pass

    paginator   = Paginator(medicines_qs, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    # Alert panel items
    all_active = list(Medicine.objects.filter(is_active=True).prefetch_related('batches'))
    low_stock_items   = sorted([m for m in all_active if m.is_low_stock],
                                key=lambda m: m.stock_quantity)[:10]
    out_of_stock_items = [m for m in all_active if m.is_out_of_stock][:10]

    expiring_batches = MedicineStock.objects.filter(
        medicine__is_active=True,
        quantity__gt=0,
        expiry_date__lte=expiry_threshold,
        expiry_date__gte=today,
    ).select_related('medicine').order_by('expiry_date')[:10]

    context = {
        'page_obj':     page_obj,
        'medicines':    page_obj.object_list,
        'paginator':    paginator,
        'per_page':     per_page,
        'per_page_options': [10, 15, 25, 50, 100],
        'search':         search,
        'category_id':    category_id,
        'status_filter':  status_filter,
        'dosage_filter':  dosage_filter,
        'highlight_pk':   highlight_pk,
        'categories':     MedicineCategory.objects.all(),
        'suppliers':      Supplier.objects.filter(is_active=True),
        'dosage_choices': Medicine.DOSAGE_FORM_CHOICES,
        'low_stock_items':    low_stock_items,
        'expiring_batches':   expiring_batches,
        'out_of_stock_items': out_of_stock_items,
        'total_medicines':     len(all_active),
        'low_stock_count':     len([m for m in all_active if m.is_low_stock]),
        'out_of_stock_count':  len([m for m in all_active if m.is_out_of_stock]),
        'expiring_soon_count': MedicineStock.objects.filter(
                                   medicine__is_active=True, quantity__gt=0,
                                   expiry_date__lte=expiry_threshold,
                                   expiry_date__gte=today).values('medicine').distinct().count(),
        'today': today,
        'can_manage': is_admin(request.user),
        'can_admin': is_admin(request.user),
    }
    return render(request, 'inventory.html', context)


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'inventory'), login_url='dashboard')
def inventory_add_view(request):
    """
    Adds a new Medicine product. The initial stock (if any) creates a
    MedicineStock batch and a StockMovement record.
    """
    if request.method == 'POST':
        try:
            barcode = request.POST.get('barcode', '').strip() or None
            if barcode and Medicine.objects.filter(barcode=barcode).exists():
                messages.error(request, f'Barcode "{barcode}" already exists.')
                return redirect('inventory')

            medicine = Medicine.objects.create(
                medicine_name    = request.POST.get('medicine_name', '').strip(),
                generic_name     = request.POST.get('generic_name', '').strip(),
                brand_name       = request.POST.get('brand_name', '').strip(),
                category_id      = request.POST.get('category') or None,
                dosage_form      = request.POST.get('dosage_form', 'tablet'),
                strength         = request.POST.get('strength', '').strip(),
                manufacturer     = request.POST.get('manufacturer', '').strip(),
                supplier_id      = request.POST.get('supplier') or None,
                barcode          = barcode,
                selling_price    = request.POST.get('selling_price') or 0,
                reorder_level    = int(request.POST.get('reorder_level') or 10),
                storage_location = request.POST.get('storage_location', '').strip(),
                is_s2_regulated  = request.POST.get('is_s2_regulated') == '1',
                created_by       = request.user,
            )

            # If initial stock provided, create the first batch
            init_qty = int(request.POST.get('stock_quantity') or 0)
            if init_qty > 0:
                batch = MedicineStock.objects.create(
                    medicine       = medicine,
                    batch_number   = request.POST.get('batch_number', '').strip(),
                    expiry_date    = request.POST.get('expiry_date') or None,
                    quantity       = init_qty,
                    purchase_price = request.POST.get('purchase_price') or 0,
                    supplier_id    = request.POST.get('supplier') or None,
                    created_by     = request.user,
                )
                StockMovement.objects.create(
                    medicine=medicine, batch=batch,
                    movement_type='in',
                    quantity=init_qty,
                    quantity_before=0, quantity_after=init_qty,
                    batch_number=batch.batch_number,
                    expiry_date=batch.expiry_date,
                    purchase_price=batch.purchase_price,
                    notes='Initial stock entry',
                    performed_by=request.user,
                )
            messages.success(request, f'Medicine "{medicine.medicine_name}" added successfully.')
        except Exception as e:
            messages.error(request, f'Error adding medicine: {e}')
    return redirect('inventory')


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'inventory'), login_url='dashboard')
def inventory_edit_view(request, pk):
    """
    Edits product-level fields only (name, dosage, price, reorder level, etc.).
    Batch details are managed via Stock In / Stock Management.
    """
    medicine = get_object_or_404(Medicine, pk=pk)
    if request.method == 'POST':
        try:
            barcode = request.POST.get('barcode', '').strip() or None
            if barcode and Medicine.objects.filter(barcode=barcode).exclude(pk=pk).exists():
                messages.error(request, f'Barcode "{barcode}" already assigned.')
                return redirect('inventory')

            medicine.medicine_name    = request.POST.get('medicine_name', '').strip()
            medicine.generic_name     = request.POST.get('generic_name', '').strip()
            medicine.brand_name       = request.POST.get('brand_name', '').strip()
            medicine.category_id      = request.POST.get('category') or None
            medicine.dosage_form      = request.POST.get('dosage_form', 'tablet')
            medicine.strength         = request.POST.get('strength', '').strip()
            medicine.manufacturer     = request.POST.get('manufacturer', '').strip()
            medicine.supplier_id      = request.POST.get('supplier') or None
            medicine.barcode          = barcode
            medicine.selling_price    = request.POST.get('selling_price') or 0
            medicine.reorder_level    = int(request.POST.get('reorder_level') or 10)
            medicine.storage_location = request.POST.get('storage_location', '').strip()
            medicine.is_s2_regulated  = request.POST.get('is_s2_regulated') == '1'
            medicine.save()
            messages.success(request, f'Medicine "{medicine.medicine_name}" updated.')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return redirect('inventory')


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def inventory_delete_view(request, pk):
    if request.method == 'POST':
        medicine = get_object_or_404(Medicine, pk=pk)
        name = medicine.medicine_name
        medicine.is_active = False
        medicine.save()
        messages.success(request, f'Medicine "{name}" removed.')
    return redirect('inventory')


@login_required(login_url='login')
def inventory_barcode_lookup(request):
    barcode = request.GET.get('barcode', '').strip()
    if not barcode:
        return JsonResponse({'found': False})
    try:
        med = Medicine.objects.get(barcode=barcode, is_active=True)
        from django.urls import reverse
        highlight_url = reverse('inventory') + f'?highlight={med.pk}&search={barcode}'
        return JsonResponse({
            'found':          True,
            'pk':             med.pk,
            'medicine_name':  med.medicine_name,
            'generic_name':   med.generic_name,
            'strength':       med.strength,
            'dosage_form':    med.dosage_form,
            'stock_quantity': med.stock_quantity,
            'stock_status':   med.stock_status,
            'highlight_url':  highlight_url,
        })
    except Medicine.DoesNotExist:
        return JsonResponse({'found': False})


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'inventory'), login_url='dashboard')
def inventory_stock_adjust(request, pk):
    """Quick +/- stock adjust from inventory page (creates a new batch for 'in')."""
    if request.method == 'POST':
        try:
            data      = json.loads(request.body)
            medicine  = get_object_or_404(Medicine, pk=pk)
            move_type = data.get('type', 'in')
            qty       = int(data.get('quantity', 0))
            notes     = data.get('notes', '')
            if qty <= 0:
                return JsonResponse({'error': 'Quantity must be > 0.'}, status=400)

            old_qty = medicine.stock_quantity

            if move_type == 'out':
                _deduct_fefo(medicine, qty, request.user,
                             movement_type='out', notes=notes)
            else:
                # Create a new batch with no expiry/batch info for quick add
                batch = MedicineStock.objects.create(
                    medicine=medicine, quantity=qty,
                    created_by=request.user,
                )
                StockMovement.objects.create(
                    medicine=medicine, batch=batch,
                    movement_type='in', quantity=qty,
                    quantity_before=old_qty, quantity_after=medicine.stock_quantity,
                    notes=notes, performed_by=request.user,
                )

            return JsonResponse({
                'success':      True,
                'new_quantity': medicine.stock_quantity,
                'stock_status': medicine.stock_status,
            })
        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid method.'}, status=405)


# ─────────────────────────────────────────────────────────────────────────────
# MEDICINE SEARCH — AJAX  (used by stock management & dispensing)
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def medicine_search_ajax(request):
    q  = request.GET.get('q', '').strip()
    qs = Medicine.objects.filter(is_active=True).prefetch_related('batches')
    if q:
        qs = qs.filter(
            Q(medicine_name__icontains=q) |
            Q(generic_name__icontains=q)  |
            Q(brand_name__icontains=q)    |
            Q(barcode__icontains=q)       |
            Q(strength__icontains=q)
        ).distinct()
    qs = qs.select_related('supplier')[:30]
    results = []
    for m in qs:
        # Earliest non-expired active batch for display
        fefo_batch = m.batches.filter(quantity__gt=0).order_by(
            F('expiry_date').asc(nulls_last=True), 'received_at'
        ).first()
        results.append({
            'pk':             m.pk,
            'is_s2_regulated': m.is_s2_regulated,
            'label':          f"{m.medicine_name} {m.strength} ({m.get_dosage_form_display()})",
            'medicine_name':  m.medicine_name,
            'generic_name':   m.generic_name,
            'strength':       m.strength,
            'dosage_form':    m.get_dosage_form_display(),
            'stock_quantity': m.stock_quantity,
            'stock_status':   m.stock_status,
            'reorder_level':  m.reorder_level,
            'supplier_id':    m.supplier_id or '',
            'supplier_name':  m.supplier.name if m.supplier else '',
            # From the FEFO batch (for display in stock management info card)
            'batch_number':   fefo_batch.batch_number if fefo_batch else '',
            'expiry_date':    str(fefo_batch.expiry_date) if fefo_batch and fefo_batch.expiry_date else '',
            'purchase_price': str(fefo_batch.purchase_price) if fefo_batch else '0.00',
            'selling_price':  str(m.selling_price),
            'barcode':        m.barcode or '',
        })
    return JsonResponse({'results': results})


# ─────────────────────────────────────────────────────────────────────────────
# STOCK MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'stock_management'), login_url='dashboard')
def stock_management_view(request):
    today = timezone.now().date()

    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    med_filter  = request.GET.get('medicine', '')
    user_filter = request.GET.get('user', '')
    type_filter = request.GET.get('type', '')
    per_page    = request.GET.get('per_page', '15')

    try:
        per_page = int(per_page)
        if per_page not in [10, 15, 25, 50, 100]:
            per_page = 15
    except ValueError:
        per_page = 15

    movements_qs = StockMovement.objects.select_related(
        'medicine', 'performed_by', 'supplier', 'batch').all()

    if date_from:
        movements_qs = movements_qs.filter(created_at__date__gte=date_from)
    if date_to:
        movements_qs = movements_qs.filter(created_at__date__lte=date_to)
    if med_filter:
        movements_qs = movements_qs.filter(medicine_id=med_filter)
    if user_filter:
        movements_qs = movements_qs.filter(performed_by_id=user_filter)
    if type_filter:
        movements_qs = movements_qs.filter(movement_type=type_filter)

    paginator   = Paginator(movements_qs, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    today_qs        = StockMovement.objects.filter(created_at__date=today)
    total_in_today  = today_qs.filter(movement_type__in=('in', 'return')).aggregate(
                          t=Sum('quantity'))['t'] or 0
    total_out_today = abs(today_qs.filter(movement_type__in=('out', 'expired', 'damaged')).aggregate(
                          t=Sum('quantity'))['t'] or 0)
    total_adj_today = today_qs.filter(movement_type__in=('adjust', 'audit')).count()
    total_mv_today  = today_qs.count()

    context = {
        'suppliers':      Supplier.objects.filter(is_active=True),
        'page_obj':       page_obj,
        'movements':      page_obj.object_list,
        'paginator':      paginator,
        'per_page':       per_page,
        'per_page_options': [10, 15, 25, 50, 100],
        'movement_types': StockMovement.MOVEMENT_TYPES,
        'date_from':      date_from,
        'date_to':        date_to,
        'med_filter':     med_filter,
        'user_filter':    user_filter,
        'type_filter':    type_filter,
        'all_medicines':  Medicine.objects.filter(is_active=True).order_by('medicine_name'),
        'all_users':      User.objects.filter(is_active=True).order_by('first_name'),
        'total_in_today':          total_in_today,
        'total_out_today':         total_out_today,
        'total_adjustments_today': total_adj_today,
        'total_movements_today':   total_mv_today,
        'today': today,
    }
    return render(request, 'stock_management.html', context)


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'stock_management'), login_url='dashboard')
def stock_in_view(request):
    """
    Records a new delivery. Always creates a new MedicineStock batch.
    If a batch with the same batch_number already exists for this medicine,
    it adds to that existing batch instead of creating a duplicate.
    """
    if request.method == 'POST':
        try:
            medicine_id    = request.POST.get('medicine')
            qty            = int(request.POST.get('quantity', 0))
            supplier_id    = request.POST.get('supplier') or None
            batch_number   = request.POST.get('batch_number', '').strip()
            expiry_date    = request.POST.get('expiry_date') or None
            purchase_price = request.POST.get('purchase_price') or 0
            reference_no   = request.POST.get('reference_no', '').strip()
            notes          = request.POST.get('notes', '').strip()

            if not medicine_id:
                messages.error(request, 'Please select a medicine.')
                return redirect('stock_management')
            if qty <= 0:
                messages.error(request, 'Quantity must be greater than zero.')
                return redirect('stock_management')

            medicine = get_object_or_404(Medicine, pk=medicine_id, is_active=True)
            old_total = medicine.stock_quantity

            # Try to match existing batch by batch_number (if provided)
            batch = None
            if batch_number:
                batch = medicine.batches.filter(batch_number=batch_number).first()

            if batch:
                # Update existing batch quantity
                batch.quantity       += qty
                batch.purchase_price  = purchase_price  # update price to latest
                if expiry_date:
                    batch.expiry_date = expiry_date
                if supplier_id:
                    batch.supplier_id = supplier_id
                batch.save()
            else:
                # Create new batch
                batch = MedicineStock.objects.create(
                    medicine       = medicine,
                    batch_number   = batch_number,
                    expiry_date    = expiry_date,
                    quantity       = qty,
                    purchase_price = purchase_price,
                    supplier_id    = supplier_id,
                    created_by     = request.user,
                )

            new_total = medicine.stock_quantity
            StockMovement.objects.create(
                medicine=medicine, batch=batch,
                movement_type='in',
                quantity=qty,
                quantity_before=old_total,
                quantity_after=new_total,
                supplier_id=supplier_id,
                batch_number=batch.batch_number,
                expiry_date=batch.expiry_date,
                purchase_price=purchase_price,
                reference_no=reference_no,
                notes=notes,
                performed_by=request.user,
            )
            messages.success(request,
                f'Stock In: +{qty} units of "{medicine.medicine_name}" '
                f'(Batch: {batch_number or "—"}). '
                f'New total: {new_total}.')
        except Exception as e:
            messages.error(request, f'Error recording Stock In: {e}')
    return redirect('stock_management')


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'stock_management'), login_url='dashboard')
def stock_out_view(request):
    """
    Records stock removal using FEFO deduction.
    """
    if request.method == 'POST':
        try:
            medicine_id     = request.POST.get('medicine')
            qty             = int(request.POST.get('quantity', 0))
            move_type       = request.POST.get('movement_type', 'out')
            dispensed_to    = request.POST.get('dispensed_to', '').strip()
            prescription_no = request.POST.get('prescription_no', '').strip()
            reference_no    = request.POST.get('reference_no', '').strip()
            notes           = request.POST.get('notes', '').strip()

            if move_type not in ('out', 'expired', 'damaged'):
                move_type = 'out'
            if not medicine_id:
                messages.error(request, 'Please select a medicine.')
                return redirect('stock_management')
            if qty <= 0:
                messages.error(request, 'Quantity must be greater than zero.')
                return redirect('stock_management')

            medicine = get_object_or_404(Medicine, pk=medicine_id, is_active=True)

            _deduct_fefo(
                medicine, qty, request.user,
                movement_type   = move_type,
                dispensed_to    = dispensed_to,
                prescription_no = prescription_no,
                notes           = notes or reference_no,
            )

            labels = {'out': 'Dispensed', 'expired': 'Expired Removal', 'damaged': 'Damaged'}
            messages.success(request,
                f'{labels.get(move_type, "Stock Out")}: -{qty} units of '
                f'"{medicine.medicine_name}". Remaining: {medicine.stock_quantity}.')
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error recording Stock Out: {e}')
    return redirect('stock_management')


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'stock_management'), login_url='dashboard')
def stock_adjust_view(request):
    """
    Manual audit adjustment — sets a specific batch to a new quantity.
    If no batch is specified, adjusts the FEFO (oldest) batch.
    """
    if request.method == 'POST':
        try:
            medicine_id  = request.POST.get('medicine')
            new_total    = int(request.POST.get('new_quantity', 0))
            move_type    = request.POST.get('movement_type', 'adjust')
            reference_no = request.POST.get('reference_no', '').strip()
            notes        = request.POST.get('notes', '').strip()
            batch_id     = request.POST.get('batch_id') or None

            if move_type not in ('adjust', 'audit', 'return'):
                move_type = 'adjust'
            if not medicine_id:
                messages.error(request, 'Please select a medicine.')
                return redirect('stock_management')
            if new_total < 0:
                messages.error(request, 'New quantity cannot be negative.')
                return redirect('stock_management')
            if not notes:
                messages.error(request, 'Please provide a reason for this adjustment.')
                return redirect('stock_management')

            medicine  = get_object_or_404(Medicine, pk=medicine_id, is_active=True)
            old_total = medicine.stock_quantity
            diff      = new_total - old_total

            if batch_id:
                batch = get_object_or_404(MedicineStock, pk=batch_id, medicine=medicine)
            else:
                # Adjust the FEFO batch
                batch = medicine.batches.filter(quantity__gt=0).order_by(
                    F('expiry_date').asc(nulls_last=True), 'received_at'
                ).first()
                if not batch:
                    # All batches empty — create one for positive adjustments
                    if diff > 0:
                        batch = MedicineStock.objects.create(
                            medicine=medicine, quantity=0, created_by=request.user
                        )
                    else:
                        messages.error(request, 'No active batches to adjust.')
                        return redirect('stock_management')

            batch.quantity = max(0, batch.quantity + diff)
            batch.save()

            StockMovement.objects.create(
                medicine=medicine, batch=batch,
                movement_type=move_type,
                quantity=diff,
                quantity_before=old_total,
                quantity_after=medicine.stock_quantity,
                batch_number=batch.batch_number,
                expiry_date=batch.expiry_date,
                reference_no=reference_no,
                notes=notes,
                performed_by=request.user,
            )
            direction = f'+{diff}' if diff >= 0 else str(diff)
            messages.success(request,
                f'Adjustment: "{medicine.medicine_name}" {old_total} → {medicine.stock_quantity} ({direction}).')
        except Exception as e:
            messages.error(request, f'Error recording adjustment: {e}')
    return redirect('stock_management')


@login_required(login_url='login')
def stock_medicine_info(request):
    pk = request.GET.get('pk', '')
    if not pk:
        return JsonResponse({'found': False})
    try:
        m = Medicine.objects.prefetch_related('batches').get(pk=pk, is_active=True)
        fefo = m.batches.filter(quantity__gt=0).order_by(
            F('expiry_date').asc(nulls_last=True), 'received_at'
        ).first()

        # All active batches for the batch list in the UI
        batches = [
            {
                'pk':             b.pk,
                'batch_number':   b.batch_number or '—',
                'expiry_date':    str(b.expiry_date) if b.expiry_date else '',
                'quantity':       b.quantity,
                'purchase_price': str(b.purchase_price),
                'supplier_id':    b.supplier_id or '',
                'is_expired':     b.is_expired,
                'is_expiring':    b.is_expiring_soon,
            }
            for b in m.batches.filter(quantity__gt=0).order_by(
                F('expiry_date').asc(nulls_last=True), 'received_at'
            )
        ]

        return JsonResponse({
            'found':          True,
            'medicine_name':  m.medicine_name,
            'generic_name':   m.generic_name,
            'strength':       m.strength,
            'dosage_form':    m.get_dosage_form_display(),
            'stock_quantity': m.stock_quantity,
            'stock_status':   m.stock_status,
            'reorder_level':  m.reorder_level,
            'supplier_id':    m.supplier_id or '',
            'selling_price':  str(m.selling_price),
            # FEFO batch details
            'batch_number':   fefo.batch_number if fefo else '',
            'expiry_date':    str(fefo.expiry_date) if fefo and fefo.expiry_date else '',
            'purchase_price': str(fefo.purchase_price) if fefo else '0.00',
            'barcode':        m.barcode or '',
            'batches':        batches,
        })
    except Medicine.DoesNotExist:
        return JsonResponse({'found': False})


# ─────────────────────────────────────────────────────────────────────────────
# SUPPLIERS
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'suppliers'), login_url='dashboard')
def suppliers_view(request):
    search        = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    per_page      = request.GET.get('per_page', '15')
    try:
        per_page = int(per_page)
        if per_page not in [10, 15, 25, 50, 100]:
            per_page = 15
    except ValueError:
        per_page = 15

    qs = Supplier.objects.prefetch_related('medicines').all()

    if search:
        qs = qs.filter(
            Q(name__icontains=search)         |
            Q(contact_name__icontains=search) |
            Q(email__icontains=search)        |
            Q(phone__icontains=search))
    if status_filter == 'active':
        qs = qs.filter(is_active=True)
    elif status_filter == 'inactive':
        qs = qs.filter(is_active=False)

    paginator   = Paginator(qs, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    context = {
        'suppliers':        page_obj.object_list,
        'page_obj':         page_obj,
        'paginator':        paginator,
        'per_page':         per_page,
        'per_page_options': [10, 15, 25, 50, 100],
        'search':                search,
        'status_filter':         status_filter,
        'total_suppliers':       Supplier.objects.count(),
        'active_suppliers':      Supplier.objects.filter(is_active=True).count(),
        'inactive_suppliers':    Supplier.objects.filter(is_active=False).count(),
        'total_products_supplied': Medicine.objects.filter(
                                       is_active=True, supplier__isnull=False).count(),
    }
    return render(request, 'suppliers.html', context)


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'suppliers'), login_url='dashboard')
def supplier_add_view(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Supplier name is required.')
            return redirect('suppliers')
        if Supplier.objects.filter(name__iexact=name).exists():
            messages.error(request, f'Supplier "{name}" already exists.')
            return redirect('suppliers')
        Supplier.objects.create(
            name         = name,
            contact_name = request.POST.get('contact_name', '').strip(),
            phone        = request.POST.get('phone', '').strip(),
            email        = request.POST.get('email', '').strip(),
            address      = request.POST.get('address', '').strip(),
            notes        = request.POST.get('notes', '').strip(),
        )
        messages.success(request, f'Supplier "{name}" added successfully.')
    return redirect('suppliers')


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'suppliers'), login_url='dashboard')
def supplier_edit_view(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Supplier name is required.')
            return redirect('suppliers')
        if Supplier.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f'Another supplier named "{name}" already exists.')
            return redirect('suppliers')
        supplier.name         = name
        supplier.contact_name = request.POST.get('contact_name', '').strip()
        supplier.phone        = request.POST.get('phone', '').strip()
        supplier.email        = request.POST.get('email', '').strip()
        supplier.address      = request.POST.get('address', '').strip()
        supplier.notes        = request.POST.get('notes', '').strip()
        supplier.is_active    = request.POST.get('is_active', 'true') == 'true'
        supplier.save()
        messages.success(request, f'Supplier "{supplier.name}" updated.')
    return redirect('suppliers')


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def supplier_delete_view(request, pk):
    if request.method == 'POST':
        supplier = get_object_or_404(Supplier, pk=pk)
        name = supplier.name
        supplier.delete()
        messages.success(request, f'Supplier "{name}" deleted.')
    return redirect('suppliers')


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'suppliers'), login_url='dashboard')
def supplier_toggle_view(request, pk):
    if request.method == 'POST':
        supplier = get_object_or_404(Supplier, pk=pk)
        supplier.is_active = not supplier.is_active
        supplier.save()
        label = 'activated' if supplier.is_active else 'deactivated'
        return JsonResponse({
            'success':   True,
            'is_active': supplier.is_active,
            'message':   f'"{supplier.name}" {label}.',
        })
    return JsonResponse({'error': 'Invalid method.'}, status=405)


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'suppliers'), login_url='dashboard')
def supplier_detail_view(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    medicines = list(
        supplier.medicines.filter(is_active=True).values_list('medicine_name', flat=True)
    )
    history_qs = StockMovement.objects.filter(
        supplier=supplier, movement_type='in'
    ).select_related('medicine').order_by('-created_at')[:20]
    purchase_history = [
        {
            'date':      mv.created_at.strftime('%b %d, %Y'),
            'medicine':  mv.medicine.medicine_name,
            'quantity':  mv.quantity,
            'reference': mv.reference_no or '',
            'notes':     mv.notes or '',
        }
        for mv in history_qs
    ]
    return JsonResponse({
        'pk':               supplier.pk,
        'name':             supplier.name,
        'contact_name':     supplier.contact_name,
        'phone':            supplier.phone,
        'email':            supplier.email,
        'address':          supplier.address,
        'notes':            supplier.notes,
        'is_active':        supplier.is_active,
        'medicines':        medicines,
        'purchase_history': purchase_history,
    })


# ─────────────────────────────────────────────────────────────────────────────
# DOCTORS
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'doctors'), login_url='dashboard')
def doctors_view(request):
    search        = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    s2_filter     = request.GET.get('s2', '')
    per_page      = request.GET.get('per_page', '15')
    try:
        per_page = int(per_page)
        if per_page not in [10, 15, 25, 50, 100]:
            per_page = 15
    except ValueError:
        per_page = 15

    qs = Doctor.objects.all()

    if search:
        qs = qs.filter(
            Q(first_name__icontains=search)   |
            Q(middle_name__icontains=search)  |
            Q(last_name__icontains=search)    |
            Q(prc_license_no__icontains=search) |
            Q(s2_license_no__icontains=search)  |
            Q(specialization__icontains=search)
        )
    if status_filter == 'active':
        qs = qs.filter(is_active=True)
    elif status_filter == 'inactive':
        qs = qs.filter(is_active=False)

    # S2 filter must be done in Python because s2_license_status is a property
    if s2_filter in ('valid', 'expiring_soon', 'expired', 'none'):
        qs = [d for d in qs if d.s2_license_status == s2_filter]

    today = timezone.now().date()

    # Summary card counts (always from full queryset, not filtered)
    all_doctors     = Doctor.objects.all()
    total_doctors   = all_doctors.count()
    active_doctors  = all_doctors.filter(is_active=True).count()
    # S2 licensed = has a license number recorded
    s2_licensed     = sum(1 for d in all_doctors if d.has_s2_license)
    # Expiring soon = S2 expiry within 30 days (valid, not expired)
    expiring_soon   = sum(1 for d in all_doctors if d.s2_is_expiring_soon)

    paginator   = Paginator(qs, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    context = {
        'doctors':          page_obj.object_list,
        'page_obj':         page_obj,
        'paginator':        paginator,
        'per_page':         per_page,
        'per_page_options': [10, 15, 25, 50, 100],
        'search':           search,
        'status_filter':    status_filter,
        's2_filter':        s2_filter,
        'total_doctors':    total_doctors,
        'active_doctors':   active_doctors,
        's2_licensed':      s2_licensed,
        'expiring_soon':    expiring_soon,
        'today':            today,
        'can_admin': is_admin(request.user),
    }
    return render(request, 'doctors.html', context)


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'doctors'), login_url='dashboard')
def doctor_add_view(request):
    if request.method == 'POST':
        prc_no = request.POST.get('prc_license_no', '').strip()
        if not prc_no:
            messages.error(request, 'PRC License No. is required.')
            return redirect('doctors')
        if Doctor.objects.filter(prc_license_no=prc_no).exists():
            messages.error(request, f'A doctor with PRC License No. "{prc_no}" already exists.')
            return redirect('doctors')

        prc_expiry = request.POST.get('prc_expiry_date', '').strip() or None
        s2_expiry  = request.POST.get('s2_expiry_date', '').strip() or None
        s2_no      = request.POST.get('s2_license_no', '').strip()

        Doctor.objects.create(
            first_name      = request.POST.get('first_name', '').strip(),
            middle_name     = request.POST.get('middle_name', '').strip(),
            last_name       = request.POST.get('last_name', '').strip(),
            specialization  = request.POST.get('specialization', '').strip(),
            prc_license_no  = prc_no,
            prc_expiry_date = prc_expiry,
            s2_license_no   = s2_no,
            s2_expiry_date  = s2_expiry,
            contact_number  = request.POST.get('contact_number', '').strip(),
            clinic_address  = request.POST.get('clinic_address', '').strip(),
        )
        messages.success(request, f'Dr. {request.POST.get("first_name", "")} '
                                   f'{request.POST.get("last_name", "")} added successfully.')
    return redirect('doctors')


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'doctors'), login_url='dashboard')
def doctor_edit_view(request, pk):
    doctor = get_object_or_404(Doctor, pk=pk)
    if request.method == 'POST':
        prc_no = request.POST.get('prc_license_no', '').strip()
        if not prc_no:
            messages.error(request, 'PRC License No. is required.')
            return redirect('doctors')
        if Doctor.objects.filter(prc_license_no=prc_no).exclude(pk=pk).exists():
            messages.error(request, f'PRC License No. "{prc_no}" is already assigned to another doctor.')
            return redirect('doctors')

        doctor.first_name      = request.POST.get('first_name', '').strip()
        doctor.middle_name     = request.POST.get('middle_name', '').strip()
        doctor.last_name       = request.POST.get('last_name', '').strip()
        doctor.specialization  = request.POST.get('specialization', '').strip()
        doctor.prc_license_no  = prc_no
        doctor.prc_expiry_date = request.POST.get('prc_expiry_date', '').strip() or None
        doctor.s2_license_no   = request.POST.get('s2_license_no', '').strip()
        doctor.s2_expiry_date  = request.POST.get('s2_expiry_date', '').strip() or None
        doctor.contact_number  = request.POST.get('contact_number', '').strip()
        doctor.clinic_address  = request.POST.get('clinic_address', '').strip()
        doctor.save()
        messages.success(request, f'Dr. {doctor.get_full_name()} updated.')
    return redirect('doctors')


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'doctors'), login_url='dashboard')
def doctor_delete_view(request, pk):
    if request.method == 'POST':
        doctor = get_object_or_404(Doctor, pk=pk)
        name   = doctor.get_full_name()
        # Soft delete — keep record for dispensing history integrity
        doctor.is_active = False
        doctor.save()
        messages.success(request, f'Dr. {name} has been deactivated.')
    return redirect('doctors')


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'doctors'), login_url='dashboard')
def doctor_toggle_view(request, pk):
    if request.method == 'POST':
        doctor = get_object_or_404(Doctor, pk=pk)
        doctor.is_active = not doctor.is_active
        doctor.save()
        label = 'activated' if doctor.is_active else 'deactivated'
        return JsonResponse({
            'success':   True,
            'is_active': doctor.is_active,
            'message':   f'Dr. {doctor.get_full_name()} {label}.',
        })
    return JsonResponse({'error': 'Invalid method.'}, status=405)


# ─────────────────────────────────────────────────────────────────────────────
# DOCTOR AJAX ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'doctors'), login_url='dashboard')
def doctor_search_ajax(request):
    """
    Typeahead search for the dispensing doctor dropdown.
    Only returns active doctors. Frontend uses this to populate the selector.
    """
    q  = request.GET.get('q', '').strip()
    qs = Doctor.objects.filter(is_active=True)
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q)  |
            Q(last_name__icontains=q)   |
            Q(prc_license_no__icontains=q) |
            Q(s2_license_no__icontains=q)  |
            Q(specialization__icontains=q)
        )
    qs = qs.order_by('last_name', 'first_name')[:30]

    results = []
    for d in qs:
        results.append({
            'pk':              d.pk,
            'label':           f"Dr. {d.get_full_name()}" +
                               (f" — {d.specialization}" if d.specialization else ''),
            'full_name':       d.get_full_name(),
            'specialization':  d.specialization,
            'prc_license_no':  d.prc_license_no,
            's2_license_no':   d.s2_license_no,
            's2_license_status': d.s2_license_status,
            's2_expiry_date':  str(d.s2_expiry_date) if d.s2_expiry_date else '',
            's2_is_valid':     d.s2_is_valid_for_dispensing,
        })
    return JsonResponse({'results': results})


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'doctors'), login_url='dashboard')
def doctor_s2_validate_ajax(request):
    """
    Given a doctor PK, returns whether their S2 license is valid for dispensing.
    Called by the dispensing frontend before allowing checkout when cart
    contains any S2-regulated medicine.

    Response shape:
        { valid: bool, reason: str, status: str, expiry: str }
    """
    pk = request.GET.get('pk', '').strip()
    if not pk:
        return JsonResponse({
            'valid':  False,
            'reason': 'No doctor selected.',
            'status': 'none',
            'expiry': '',
        })
    try:
        doctor = Doctor.objects.get(pk=pk, is_active=True)
    except Doctor.DoesNotExist:
        return JsonResponse({
            'valid':  False,
            'reason': 'Doctor not found or is inactive.',
            'status': 'none',
            'expiry': '',
        })

    status = doctor.s2_license_status
    expiry = str(doctor.s2_expiry_date) if doctor.s2_expiry_date else ''

    reason_map = {
        'none':          f'Dr. {doctor.get_full_name()} has no S2 license on record.',
        'expired':       f'Dr. {doctor.get_full_name()}\'s S2 license expired on {expiry}.',
        'expiring_soon': None,   # valid but frontend can show a warning
        'valid':         None,
    }

    is_valid = doctor.s2_is_valid_for_dispensing
    reason   = reason_map.get(status, 'Unknown status.')

    return JsonResponse({
        'valid':  is_valid,
        'reason': reason or '',
        'status': status,
        'expiry': expiry,
        'name':   doctor.get_full_name(),
    })



# ─────────────────────────────────────────────────────────────────────────────
# USER MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def user_management_view(request):
    search_query = request.GET.get('search', '').strip()
    users = User.objects.select_related(
        'profile', 'permissions'
    ).order_by('-date_joined')

    if search_query:
        users = users.filter(
            Q(username__icontains=search_query)   |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)  |
            Q(email__icontains=search_query)
        )

    # Ensure every user has a permissions row (handles users created before this feature)
    for u in users:
        UserPermission.objects.get_or_create(user=u)

    # Re-fetch after ensuring rows exist so template can read permissions
    users = User.objects.select_related(
        'profile', 'permissions'
    ).order_by('-date_joined')

    if search_query:
        users = users.filter(
            Q(username__icontains=search_query)   |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)  |
            Q(email__icontains=search_query)
        )

    context = {
        'users':          users,
        'search_query':   search_query,
        'total_users':    User.objects.count(),
        'active_users':   User.objects.filter(is_active=True).count(),
        'inactive_users': User.objects.filter(is_active=False).count(),
        'module_choices': UserPermission.MODULE_CHOICES,
    }
    return render(request, 'user_management.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def user_add_view(request):
    if request.method == 'POST':
        username    = request.POST.get('username', '').strip()
        first_name  = request.POST.get('first_name', '').strip()
        middle_name = request.POST.get('middle_name', '').strip()
        last_name   = request.POST.get('last_name', '').strip()
        email       = request.POST.get('email', '').strip()
        position    = request.POST.get('position', '').strip()
        password    = request.POST.get('password', '')
        confirm_pw  = request.POST.get('confirm_password', '')
        role        = request.POST.get('role', 'user')
        if User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" is taken.')
        elif not password:
            messages.error(request, 'Password cannot be empty.')
        elif password != confirm_pw:
            messages.error(request, 'Passwords do not match.')
        elif len(password) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
        else:
            user = User.objects.create_user(
                username=username, first_name=first_name,
                last_name=last_name, email=email, password=password,
                is_staff=role in ('staff', 'admin'),
            )
            user.profile.middle_name = middle_name
            user.profile.position = position
            user.profile.role = role
            user.profile.save()

            # Save module permissions (skipped for admins)
            if role != 'admin' and not user.is_superuser:
                perm, _ = UserPermission.objects.get_or_create(user=user)
                for module_key, _ in UserPermission.MODULE_CHOICES:
                    setattr(perm, module_key, request.POST.get(f'perm_{module_key}') == 'on')
                perm.save()

            messages.success(request, f'User "{user.username}" created.')
    return redirect('user_management')


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def user_edit_view(request, pk):
    user_obj = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user_obj.first_name = request.POST.get('first_name', '').strip()
        user_obj.last_name  = request.POST.get('last_name', '').strip()
        user_obj.email      = request.POST.get('email', '').strip()
        user_obj.is_active  = request.POST.get('is_active') == 'on'
        role = request.POST.get('role', 'user')
        user_obj.is_staff = role in ('staff', 'admin')
        new_pw = request.POST.get('password', '')
        if new_pw:
            if len(new_pw) < 8:
                messages.error(request, 'Password must be at least 8 characters.')
                return redirect('user_management')
            user_obj.set_password(new_pw)
        user_obj.save()
        user_obj.profile.middle_name = request.POST.get('middle_name', '').strip()
        user_obj.profile.position = request.POST.get('position', '').strip()
        user_obj.profile.role = role
        user_obj.profile.save()

        # ── Save module permissions (skipped for admins — they always get everything)
        if not user_obj.is_superuser and get_role(user_obj) != 'admin':
            perm, _ = UserPermission.objects.get_or_create(user=user_obj)
            for module_key, _ in UserPermission.MODULE_CHOICES:
                setattr(perm, module_key, request.POST.get(f'perm_{module_key}') == 'on')
            perm.save()

        messages.success(request, f'User "{user_obj.username}" updated.')
    return redirect('user_management')


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def user_delete_view(request, pk):
    if request.method == 'POST':
        user_obj = get_object_or_404(User, pk=pk)
        if user_obj == request.user:
            messages.error(request, 'Cannot delete your own account.')
        else:
            username = user_obj.username
            user_obj.delete()
            messages.success(request, f'User "{username}" deleted.')
    return redirect('user_management')


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def user_toggle_active_view(request, pk):
    if request.method == 'POST':
        user_obj = get_object_or_404(User, pk=pk)
        if user_obj == request.user:
            return JsonResponse({'error': 'Cannot deactivate your own account.'}, status=400)
        user_obj.is_active = not user_obj.is_active
        user_obj.save()
        label = 'activated' if user_obj.is_active else 'deactivated'
        return JsonResponse({
            'success':   True,
            'is_active': user_obj.is_active,
            'message':   f'{user_obj.get_full_name() or user_obj.username} {label}.',
        })
    return JsonResponse({'error': 'Invalid method.'}, status=405)


# ─────────────────────────────────────────────────────────────────────────────
# SETTINGS
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def settings_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_profile':
            request.user.first_name = request.POST.get('first_name', '').strip()
            request.user.last_name  = request.POST.get('last_name', '').strip()
            request.user.email      = request.POST.get('email', '').strip()
            request.user.save()
            request.user.profile.middle_name = request.POST.get('middle_name', '').strip()
            request.user.profile.position    = request.POST.get('position', '').strip()
            request.user.profile.save()
            messages.success(request, 'Profile updated.')
        elif action == 'change_password':
            cur = request.POST.get('current_password', '')
            new = request.POST.get('new_password', '')
            cfm = request.POST.get('confirm_password', '')
            if not request.user.check_password(cur):
                messages.error(request, 'Current password incorrect.')
            elif new != cfm:
                messages.error(request, 'New passwords do not match.')
            elif len(new) < 8:
                messages.error(request, 'Password must be at least 8 characters.')
            else:
                request.user.set_password(new)
                request.user.save()
                updated = authenticate(request, username=request.user.username, password=new)
                if updated:
                    login(request, updated)
                messages.success(request, 'Password changed.')
        return redirect('settings')
    return render(request, 'settings.html', {'user_obj': request.user})


# ─────────────────────────────────────────────────────────────────────────────
# REPORTS
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'reports'), login_url='dashboard')
def reports_view(request):
    today              = timezone.now().date()
    expiry_threshold   = today + timedelta(days=90)
    date_from, date_to = _parse_date_range(request, default_days=30)
    active_tab         = request.GET.get('tab', 'inventory')

    all_active = Medicine.objects.filter(is_active=True).prefetch_related('batches')

    total_medicines     = all_active.count()
    low_stock_count     = len([m for m in all_active if m.is_low_stock])
    out_of_stock_count  = len([m for m in all_active if m.is_out_of_stock])
    expired_count       = MedicineStock.objects.filter(
                              medicine__is_active=True, quantity__gt=0,
                              expiry_date__lt=today).count()
    expiring_soon_count = MedicineStock.objects.filter(
                              medicine__is_active=True, quantity__gt=0,
                              expiry_date__gte=today,
                              expiry_date__lte=expiry_threshold).count()

    stock_value = MedicineStock.objects.filter(
        medicine__is_active=True, quantity__gt=0
    ).annotate(
        line_val=ExpressionWrapper(
            F('medicine__selling_price') * F('quantity'),
            output_field=DecimalField()
        )
    ).aggregate(total=Sum('line_val'))['total'] or 0

    current_stock_qs = all_active.select_related('category', 'supplier').order_by('medicine_name')
    low_stock_qs     = [m for m in current_stock_qs if m.is_low_stock]
    expired_qs       = MedicineStock.objects.filter(
                           medicine__is_active=True, quantity__gt=0,
                           expiry_date__lt=today
                       ).select_related('medicine__category', 'medicine__supplier').order_by('expiry_date')
    expiring_qs      = MedicineStock.objects.filter(
                           medicine__is_active=True, quantity__gt=0,
                           expiry_date__gte=today, expiry_date__lte=expiry_threshold
                       ).select_related('medicine__category', 'medicine__supplier').order_by('expiry_date')

    category_stock = list(
        MedicineStock.objects.filter(medicine__is_active=True, quantity__gt=0)
        .values('medicine__category__name')
        .annotate(total=Sum('quantity'))
        .order_by('-total')[:10]
    )
    cat_labels = json.dumps([x['medicine__category__name'] or 'Uncategorized' for x in category_stock])
    cat_values = json.dumps([x['total'] for x in category_stock])

    sales_qs = StockMovement.objects.filter(
        movement_type='out',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).select_related('medicine', 'performed_by')

    daily_sales = list(
        sales_qs.annotate(day=TruncDay('created_at'))
        .values('day')
        .annotate(units=Sum(ExpressionWrapper(-F('quantity'), output_field=FloatField())))
        .order_by('day')
    )
    sales_labels = json.dumps([str(d['day'].date()) for d in daily_sales])
    sales_units  = json.dumps([d['units'] for d in daily_sales])

    sales_with_value = sales_qs.annotate(
        rev=ExpressionWrapper(
            -F('quantity') * F('medicine__selling_price'),
            output_field=DecimalField()
        )
    )
    total_revenue      = sales_with_value.aggregate(t=Sum('rev'))['t'] or 0
    total_units_sold   = sales_qs.aggregate(
        t=Sum(ExpressionWrapper(-F('quantity'), output_field=FloatField()))
    )['t'] or 0
    total_transactions = sales_qs.count()

    top_medicines = list(
        sales_qs.values('medicine__medicine_name', 'medicine__strength')
        .annotate(
            units=Sum(ExpressionWrapper(-F('quantity'), output_field=FloatField())),
            revenue=Sum(ExpressionWrapper(
                -F('quantity') * F('medicine__selling_price'), output_field=DecimalField()
            ))
        ).order_by('-units')[:10]
    )
    top_med_labels   = json.dumps([f"{m['medicine__medicine_name']} {m['medicine__strength']}" for m in top_medicines])
    top_med_units    = json.dumps([m['units'] for m in top_medicines])
    top_med_revenues = json.dumps([float(m['revenue'] or 0) for m in top_medicines])

    sales_detail = list(
        sales_qs.annotate(
            rev=ExpressionWrapper(
                -F('quantity') * F('medicine__selling_price'), output_field=DecimalField()
            )
        ).values(
            'created_at', 'medicine__medicine_name', 'medicine__strength',
            'quantity', 'dispensed_to', 'prescription_no', 'reference_no',
            'rev', 'performed_by__first_name', 'performed_by__last_name',
            'performed_by__username'
        ).order_by('-created_at')[:200]
    )

    # COGS: use purchase_price recorded on StockMovement at time of deduction
    cogs_qs = StockMovement.objects.filter(
        movement_type='in',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
        purchase_price__isnull=False,
    ).annotate(
        cost=ExpressionWrapper(F('quantity') * F('purchase_price'), output_field=DecimalField())
    )
    total_cogs    = cogs_qs.aggregate(t=Sum('cost'))['t'] or 0
    gross_profit  = float(total_revenue) - float(total_cogs)
    profit_margin = (gross_profit / float(total_revenue) * 100) if total_revenue else 0

    monthly_rev = list(
        StockMovement.objects.filter(movement_type='out')
        .annotate(
            month=TruncMonth('created_at'),
            rev=ExpressionWrapper(-F('quantity') * F('medicine__selling_price'), output_field=DecimalField())
        ).values('month').annotate(total_rev=Sum('rev')).order_by('month')
        .filter(month__date__gte=today - timedelta(days=180))
    )
    monthly_rev_labels = json.dumps([str(m['month'].date()) for m in monthly_rev])
    monthly_rev_values = json.dumps([float(m['total_rev'] or 0) for m in monthly_rev])

    monthly_cost = list(
        StockMovement.objects.filter(movement_type='in', purchase_price__isnull=False)
        .annotate(
            month=TruncMonth('created_at'),
            cost=ExpressionWrapper(F('quantity') * F('purchase_price'), output_field=DecimalField())
        ).values('month').annotate(total_cost=Sum('cost')).order_by('month')
        .filter(month__date__gte=today - timedelta(days=180))
    )
    monthly_cost_labels = json.dumps([str(m['month'].date()) for m in monthly_cost])
    monthly_cost_values = json.dumps([float(m['total_cost'] or 0) for m in monthly_cost])

    profit_by_medicine = list(
        StockMovement.objects.filter(
            movement_type='out',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        ).values('medicine__medicine_name', 'medicine__strength')
        .annotate(
            units_sold=Sum(ExpressionWrapper(-F('quantity'), output_field=FloatField())),
            revenue=Sum(ExpressionWrapper(
                -F('quantity') * F('medicine__selling_price'), output_field=DecimalField()
            )),
            cogs=Sum(ExpressionWrapper(
                -F('quantity') * F('purchase_price'), output_field=DecimalField()
            )),
        ).order_by('-revenue')[:15]
    )
    for row in profit_by_medicine:
        rev  = float(row['revenue'] or 0)
        cost = float(row['cogs'] or 0)
        row['gross_profit'] = rev - cost
        row['margin_pct']   = ((rev - cost) / rev * 100) if rev else 0

    context = {
        'active_tab': active_tab,
        'date_from':  date_from,
        'date_to':    date_to,
        'today':      today,
        'total_medicines': total_medicines, 'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count, 'expired_count': expired_count,
        'expiring_soon_count': expiring_soon_count, 'stock_value': stock_value,
        'current_stock_qs': current_stock_qs, 'low_stock_qs': low_stock_qs,
        'expired_qs': expired_qs, 'expiring_qs': expiring_qs,
        'cat_labels': cat_labels, 'cat_values': cat_values,
        'total_revenue': total_revenue, 'total_units_sold': total_units_sold,
        'total_transactions': total_transactions, 'sales_detail': sales_detail,
        'top_medicines': top_medicines, 'sales_labels': sales_labels,
        'sales_units': sales_units, 'top_med_labels': top_med_labels,
        'top_med_units': top_med_units, 'top_med_revenues': top_med_revenues,
        'total_cogs': total_cogs, 'gross_profit': gross_profit,
        'profit_margin': round(profit_margin, 1), 'profit_by_medicine': profit_by_medicine,
        'monthly_rev_labels': monthly_rev_labels, 'monthly_rev_values': monthly_rev_values,
        'monthly_cost_labels': monthly_cost_labels, 'monthly_cost_values': monthly_cost_values,
    }
    return render(request, 'reports.html', context)


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT — EXCEL  (unchanged structure, batch-aware expiry data)
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'reports'), login_url='dashboard')
def report_export_excel(request):
    if not OPENPYXL_AVAILABLE:
        messages.error(request, 'openpyxl is not installed. Run: pip install openpyxl')
        return redirect('reports')

    report_type        = request.GET.get('type', 'current_stock')
    date_from, date_to = _parse_date_range(request)
    today              = timezone.now().date()
    expiry_threshold   = today + timedelta(days=90)

    wb = openpyxl.Workbook()
    ws = wb.active

    header_fill  = PatternFill('solid', fgColor='00ADB5')
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    title_font   = Font(bold=True, size=14, color='0D1B2A')
    sub_font     = Font(size=9, color='78909C')
    border_side  = Side(style='thin', color='E0E6ED')
    cell_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left',   vertical='center')

    def style_header_row(row_num, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = center_align; cell.border = cell_border

    def style_data_row(row_num, col_count, alternate=False):
        alt_fill = PatternFill('solid', fgColor='F8FAFC' if alternate else 'FFFFFF')
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = alt_fill; cell.border = cell_border; cell.alignment = left_align

    def write_report_title(title, subtitle=''):
        ws.merge_cells(f'A1:{get_column_letter(10)}1')
        ws['A1'].value = 'MHARSMC — Electronic Pharmacy Inventory System'
        ws['A1'].font  = Font(bold=True, size=16, color='0D1B2A')
        ws['A1'].alignment = center_align
        ws.merge_cells(f'A2:{get_column_letter(10)}2')
        ws['A2'].value = title; ws['A2'].font = title_font; ws['A2'].alignment = center_align
        ws.merge_cells(f'A3:{get_column_letter(10)}3')
        ws['A3'].value = subtitle or f'Generated: {timezone.now().strftime("%B %d, %Y %I:%M %p")}'
        ws['A3'].font  = sub_font; ws['A3'].alignment = center_align
        return 5

    if report_type == 'current_stock':
        ws.title  = 'Current Stock'
        start_row = write_report_title('Current Stock Report')
        headers   = ['#', 'Medicine Name', 'Generic Name', 'Category',
                     'Dosage Form', 'Strength', 'Supplier',
                     'Stock Qty', 'Reorder Level', 'Status']
        ws.append([])
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col).value = h
        style_header_row(start_row, len(headers))
        qs = Medicine.objects.filter(is_active=True).select_related(
            'category', 'supplier').prefetch_related('batches').order_by('medicine_name')
        for i, m in enumerate(qs, 1):
            r = start_row + i
            status = 'Out of Stock' if m.is_out_of_stock else \
                     'Low Stock'    if m.is_low_stock    else 'In Stock'
            ws.append([i, m.medicine_name, m.generic_name,
                        m.category.name if m.category else '—',
                        m.get_dosage_form_display(), m.strength,
                        m.supplier.name if m.supplier else '—',
                        m.stock_quantity, m.reorder_level, status])
            style_data_row(r, len(headers), i % 2 == 0)
            status_cell = ws.cell(row=r, column=10)
            status_cell.font = Font(bold=True, color='DC2626' if m.is_out_of_stock else
                                    ('D97706' if m.is_low_stock else '16A34A'))
        col_widths = [5, 28, 25, 18, 14, 12, 22, 10, 12, 12]

    elif report_type == 'low_stock':
        ws.title  = 'Low Stock'
        start_row = write_report_title('Low Stock Report', f'Medicines at or below reorder level — {today}')
        headers   = ['#', 'Medicine Name', 'Generic Name', 'Category',
                     'Strength', 'Stock Qty', 'Reorder Level', 'Deficit', 'Supplier']
        ws.append([])
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col).value = h
        style_header_row(start_row, len(headers))
        qs = [m for m in Medicine.objects.filter(is_active=True).select_related(
              'category', 'supplier').prefetch_related('batches').order_by('medicine_name')
              if m.is_low_stock]
        for i, m in enumerate(qs, 1):
            r = start_row + i
            ws.append([i, m.medicine_name, m.generic_name,
                        m.category.name if m.category else '—',
                        m.strength, m.stock_quantity, m.reorder_level,
                        m.reorder_level - m.stock_quantity,
                        m.supplier.name if m.supplier else '—'])
            style_data_row(r, len(headers), i % 2 == 0)
        col_widths = [5, 28, 25, 18, 12, 10, 12, 10, 22]

    elif report_type == 'expired':
        ws.title  = 'Expired Batches'
        start_row = write_report_title('Expired Batches Report', f'As of {today}')
        headers   = ['#', 'Medicine Name', 'Generic Name', 'Category',
                     'Strength', 'Batch Number', 'Expiry Date', 'Qty Remaining', 'Supplier']
        ws.append([])
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col).value = h
        style_header_row(start_row, len(headers))
        qs = MedicineStock.objects.filter(
            medicine__is_active=True, quantity__gt=0, expiry_date__lt=today
        ).select_related('medicine__category', 'medicine__supplier').order_by('expiry_date')
        for i, b in enumerate(qs, 1):
            r = start_row + i
            m = b.medicine
            ws.append([i, m.medicine_name, m.generic_name,
                        m.category.name if m.category else '—',
                        m.strength, b.batch_number or '—',
                        str(b.expiry_date), b.quantity,
                        b.supplier.name if b.supplier else (m.supplier.name if m.supplier else '—')])
            style_data_row(r, len(headers), i % 2 == 0)
            ws.cell(row=r, column=7).font = Font(bold=True, color='DC2626')
        col_widths = [5, 28, 25, 18, 12, 16, 14, 10, 22]

    elif report_type == 'expiring':
        ws.title  = 'Expiring Soon'
        start_row = write_report_title('Expiring Soon Report', f'Batches expiring within 90 days — {today}')
        headers   = ['#', 'Medicine Name', 'Generic Name', 'Category',
                     'Strength', 'Batch Number', 'Expiry Date', 'Days Left', 'Qty', 'Supplier']
        ws.append([])
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col).value = h
        style_header_row(start_row, len(headers))
        qs = MedicineStock.objects.filter(
            medicine__is_active=True, quantity__gt=0,
            expiry_date__gte=today, expiry_date__lte=expiry_threshold
        ).select_related('medicine__category', 'medicine__supplier').order_by('expiry_date')
        for i, b in enumerate(qs, 1):
            r         = start_row + i
            m         = b.medicine
            days_left = (b.expiry_date - today).days
            ws.append([i, m.medicine_name, m.generic_name,
                        m.category.name if m.category else '—',
                        m.strength, b.batch_number or '—',
                        str(b.expiry_date), days_left, b.quantity,
                        b.supplier.name if b.supplier else (m.supplier.name if m.supplier else '—')])
            style_data_row(r, len(headers), i % 2 == 0)
            ws.cell(row=r, column=8).font = Font(
                bold=True, color='D97706' if days_left <= 30 else '1A2433'
            )
        col_widths = [5, 28, 25, 18, 12, 16, 14, 10, 10, 22]

    elif report_type == 'sales':
        ws.title  = 'Sales Report'
        start_row = write_report_title('Sales Report', f'{date_from} to {date_to}')
        headers   = ['#', 'Date', 'Medicine', 'Strength', 'Qty Dispensed',
                     'Selling Price', 'Revenue', 'Dispensed To', 'Prescription No.', 'Performed By']
        ws.append([])
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col).value = h
        style_header_row(start_row, len(headers))
        qs = StockMovement.objects.filter(
            movement_type='out',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        ).select_related('medicine', 'performed_by').order_by('-created_at')
        for i, mv in enumerate(qs, 1):
            r   = start_row + i
            qty = abs(mv.quantity)
            rev = qty * mv.medicine.selling_price
            by  = mv.performed_by.get_full_name() if mv.performed_by else '—'
            ws.append([i, mv.created_at.strftime('%Y-%m-%d %H:%M'),
                        mv.medicine.medicine_name, mv.medicine.strength,
                        qty, float(mv.medicine.selling_price), float(rev),
                        mv.dispensed_to or '—', mv.prescription_no or '—', by])
            style_data_row(r, len(headers), i % 2 == 0)
            ws.cell(row=r, column=7).number_format = '#,##0.00'
            ws.cell(row=r, column=6).number_format = '#,##0.00'
        col_widths = [5, 18, 28, 12, 14, 13, 13, 22, 16, 20]

    elif report_type == 'financial':
        ws.title  = 'Financial Report'
        start_row = write_report_title('Financial Report', f'{date_from} to {date_to}')
        headers   = ['#', 'Medicine', 'Strength', 'Units Sold',
                     'Selling Price', 'Purchase Price', 'Revenue', 'COGS', 'Gross Profit', 'Margin %']
        ws.append([])
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col).value = h
        style_header_row(start_row, len(headers))
        qs = StockMovement.objects.filter(
            movement_type='out',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        ).values('medicine__medicine_name', 'medicine__strength',
                 'medicine__selling_price', 'medicine__purchase_price') \
         .annotate(
             units=Sum(ExpressionWrapper(-F('quantity'), output_field=FloatField())),
             rev=Sum(ExpressionWrapper(
                 -F('quantity') * F('medicine__selling_price'), output_field=DecimalField()
             )),
             cost=Sum(ExpressionWrapper(
                 -F('quantity') * F('medicine__purchase_price'), output_field=DecimalField()
             )),
         ).order_by('-rev')
        for i, row in enumerate(qs, 1):
            r      = start_row + i
            rev    = float(row['rev'] or 0)
            cost   = float(row['cost'] or 0)
            profit = rev - cost
            margin = (profit / rev * 100) if rev else 0
            ws.append([i, row['medicine__medicine_name'], row['medicine__strength'],
                        row['units'], float(row['medicine__selling_price']),
                        float(row['medicine__purchase_price']),
                        rev, cost, profit, round(margin, 1)])
            style_data_row(r, len(headers), i % 2 == 0)
            for c in [5, 6, 7, 8, 9]:
                ws.cell(row=r, column=c).number_format = '#,##0.00'
            ws.cell(row=r, column=9).font = Font(
                bold=True, color='16A34A' if profit >= 0 else 'DC2626'
            )
        col_widths = [5, 28, 12, 12, 13, 13, 13, 13, 13, 10]
    else:
        col_widths = [20] * 5

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=6, column=1)

    filename = f'MHARSMC_{report_type}_{today}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT — PDF
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'reports'), login_url='dashboard')
def report_export_pdf(request):
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'reportlab is not installed. Run: pip install reportlab')
        return redirect('reports')

    report_type        = request.GET.get('type', 'current_stock')
    date_from, date_to = _parse_date_range(request)
    today              = timezone.now().date()
    expiry_threshold   = today + timedelta(days=90)

    buffer = io.BytesIO()
    FOLIO  = (8.5 * inch, 13 * inch)
    wide_types = {'current_stock', 'sales', 'financial', 'expiring', 'expired', 'low_stock'}
    pagesize   = landscape(FOLIO) if report_type in wide_types else FOLIO

    doc = SimpleDocTemplate(buffer, pagesize=pagesize,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles   = getSampleStyleSheet()
    elements = []

    TEAL    = colors.HexColor('#00ADB5')
    DARK    = colors.HexColor('#0D1B2A')
    MUTED   = colors.HexColor('#78909C')
    RED     = colors.HexColor('#DC2626')
    ORANGE  = colors.HexColor('#D97706')
    GREEN   = colors.HexColor('#16A34A')
    ALT_ROW = colors.HexColor('#EEF2F7')

    title_style = ParagraphStyle('Title', parent=styles['Normal'],
                                 fontSize=12, textColor=DARK,
                                 fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=2)
    sub_style   = ParagraphStyle('Sub', parent=styles['Normal'],
                                 fontSize=9, textColor=MUTED, alignment=TA_CENTER, spaceAfter=4)
    body_style  = ParagraphStyle('Body', parent=styles['Normal'], fontSize=8, textColor=DARK)

    def make_table_style(num_rows, num_cols):
        return TableStyle([
            ('BACKGROUND',  (0,0),  (-1,0),  TEAL),
            ('TEXTCOLOR',   (0,0),  (-1,0),  colors.white),
            ('FONTNAME',    (0,0),  (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',    (0,0),  (-1,0),  8),
            ('ALIGN',       (0,0),  (-1,-1), 'CENTER'),
            ('VALIGN',      (0,0),  (-1,-1), 'MIDDLE'),
            ('FONTNAME',    (0,1),  (-1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,1),  (-1,-1), 7.5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, ALT_ROW]),
            ('GRID',        (0,0),  (-1,-1), 0.3, colors.HexColor('#E0E6ED')),
            ('BOTTOMPADDING',(0,0), (-1,-1), 4),
            ('TOPPADDING',  (0,0),  (-1,-1), 4),
            ('LEFTPADDING', (0,0),  (-1,-1), 5),
            ('RIGHTPADDING',(0,0),  (-1,-1), 5),
        ])

    elements.append(Paragraph('MHARSMC — Electronic Pharmacy Inventory System', title_style))
    elements.append(HRFlowable(width='100%', thickness=2, color=TEAL, spaceAfter=6))

    report_titles = {
        'current_stock': 'Current Stock Report',
        'low_stock':     'Low Stock Report',
        'expired':       'Expired Batches Report',
        'expiring':      'Expiring Soon Report',
        'sales':         f'Sales Report: {date_from} to {date_to}',
        'financial':     f'Financial Report: {date_from} to {date_to}',
    }
    elements.append(Paragraph(report_titles.get(report_type, 'Report'), title_style))
    elements.append(Paragraph(
        f'Generated: {timezone.now().strftime("%B %d, %Y %I:%M %p")} | '
        f'Mayor Hilarion A. Ramiro Sr. Medical Center', sub_style
    ))
    elements.append(Spacer(1, 0.4*cm))

    if report_type == 'current_stock':
        qs   = Medicine.objects.filter(is_active=True).select_related(
                   'category', 'supplier').prefetch_related('batches').order_by('medicine_name')
        data = [['#', 'Medicine Name', 'Generic', 'Category',
                 'Form', 'Strength', 'Stock', 'Reorder', 'Status']]
        ts   = make_table_style(qs.count() + 1, 9)
        for i, m in enumerate(qs, 1):
            status = 'Out' if m.is_out_of_stock else 'Low' if m.is_low_stock else 'OK'
            data.append([i, m.medicine_name, m.generic_name,
                          m.category.name if m.category else '—',
                          m.get_dosage_form_display(), m.strength,
                          m.stock_quantity, m.reorder_level, status])
            color = RED if m.is_out_of_stock else (ORANGE if m.is_low_stock else GREEN)
            ts.add('TEXTCOLOR', (8, i), (8, i), color)
            ts.add('FONTNAME',  (8, i), (8, i), 'Helvetica-Bold')
        col_widths = [1*cm, 5.5*cm, 4.5*cm, 3*cm, 2.5*cm, 2.5*cm, 1.8*cm, 2*cm, 1.8*cm]

    elif report_type == 'low_stock':
        qs   = [m for m in Medicine.objects.filter(is_active=True).select_related(
                    'category', 'supplier').prefetch_related('batches')
                if m.is_low_stock]
        data = [['#', 'Medicine', 'Generic', 'Category',
                 'Strength', 'Stock', 'Reorder', 'Deficit', 'Supplier']]
        ts   = make_table_style(len(qs) + 1, 9)
        for i, m in enumerate(qs, 1):
            data.append([i, m.medicine_name, m.generic_name,
                          m.category.name if m.category else '—',
                          m.strength, m.stock_quantity, m.reorder_level,
                          m.reorder_level - m.stock_quantity,
                          m.supplier.name if m.supplier else '—'])
        col_widths = [1*cm, 5.5*cm, 4.5*cm, 3*cm, 2.5*cm, 1.8*cm, 2*cm, 2*cm, 4*cm]

    elif report_type == 'expired':
        qs   = MedicineStock.objects.filter(
                   medicine__is_active=True, quantity__gt=0, expiry_date__lt=today
               ).select_related('medicine__category', 'medicine__supplier').order_by('expiry_date')
        data = [['#', 'Medicine', 'Generic', 'Category',
                 'Strength', 'Batch', 'Expiry Date', 'Qty', 'Supplier']]
        ts   = make_table_style(qs.count() + 1, 9)
        for i, b in enumerate(qs, 1):
            m = b.medicine
            data.append([i, m.medicine_name, m.generic_name,
                          m.category.name if m.category else '—',
                          m.strength, b.batch_number or '—',
                          str(b.expiry_date), b.quantity,
                          b.supplier.name if b.supplier else (m.supplier.name if m.supplier else '—')])
            ts.add('TEXTCOLOR', (6, i), (6, i), RED)
            ts.add('FONTNAME',  (6, i), (6, i), 'Helvetica-Bold')
        col_widths = [1*cm, 5.5*cm, 4.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.8*cm, 1.5*cm, 4*cm]

    elif report_type == 'expiring':
        qs   = MedicineStock.objects.filter(
                   medicine__is_active=True, quantity__gt=0,
                   expiry_date__gte=today, expiry_date__lte=expiry_threshold
               ).select_related('medicine__category', 'medicine__supplier').order_by('expiry_date')
        data = [['#', 'Medicine', 'Generic', 'Category',
                 'Strength', 'Batch', 'Expiry Date', 'Days Left', 'Qty']]
        ts   = make_table_style(qs.count() + 1, 9)
        for i, b in enumerate(qs, 1):
            m         = b.medicine
            days_left = (b.expiry_date - today).days
            data.append([i, m.medicine_name, m.generic_name,
                          m.category.name if m.category else '—',
                          m.strength, b.batch_number or '—',
                          str(b.expiry_date), days_left, b.quantity])
            color = RED if days_left <= 30 else ORANGE
            ts.add('TEXTCOLOR', (7, i), (7, i), color)
            ts.add('FONTNAME',  (7, i), (7, i), 'Helvetica-Bold')
        col_widths = [1*cm, 5.5*cm, 4.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.8*cm, 2*cm, 2*cm]

    elif report_type == 'sales':
        qs   = StockMovement.objects.filter(
                   movement_type='out',
                   created_at__date__gte=date_from,
                   created_at__date__lte=date_to,
               ).select_related('medicine', 'performed_by').order_by('-created_at')
        data = [['#', 'Date', 'Medicine', 'Strength',
                 'Qty', 'Sell Price', 'Revenue', 'Dispensed To', 'By']]
        ts        = make_table_style(qs.count() + 1, 9)
        total_rev = 0
        for i, mv in enumerate(qs, 1):
            qty = abs(mv.quantity)
            rev = float(qty * mv.medicine.selling_price)
            total_rev += rev
            by  = mv.performed_by.get_full_name() if mv.performed_by else '—'
            data.append([i, mv.created_at.strftime('%m/%d/%Y'),
                          mv.medicine.medicine_name, mv.medicine.strength,
                          qty, f'₱{mv.medicine.selling_price:,.2f}',
                          f'₱{rev:,.2f}', mv.dispensed_to or '—', by])
        data.append(['', '', '', '', '', 'TOTAL', f'₱{total_rev:,.2f}', '', ''])
        ts.add('FONTNAME',   (0, len(data)-1), (-1, len(data)-1), 'Helvetica-Bold')
        ts.add('BACKGROUND', (0, len(data)-1), (-1, len(data)-1), ALT_ROW)
        ts.add('TEXTCOLOR',  (6, len(data)-1), (6,  len(data)-1), GREEN)
        col_widths = [1*cm, 2.5*cm, 5.5*cm, 2.5*cm, 1.5*cm, 2.5*cm, 2.5*cm, 4*cm, 3.5*cm]

    elif report_type == 'financial':
        qs = StockMovement.objects.filter(
            movement_type='out',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        ).values('medicine__medicine_name', 'medicine__strength',
                 'medicine__selling_price', 'medicine__purchase_price') \
         .annotate(
             units=Sum(ExpressionWrapper(-F('quantity'), output_field=FloatField())),
             rev=Sum(ExpressionWrapper(
                 -F('quantity') * F('medicine__selling_price'), output_field=DecimalField()
             )),
             cost=Sum(ExpressionWrapper(
                 -F('quantity') * F('medicine__purchase_price'), output_field=DecimalField()
             )),
         ).order_by('-rev')
        data = [['#', 'Medicine', 'Strength', 'Units',
                 'Revenue', 'COGS', 'Gross Profit', 'Margin %']]
        ts   = make_table_style(qs.count() + 1, 8)
        for i, row in enumerate(qs, 1):
            rev    = float(row['rev'] or 0)
            cost   = float(row['cost'] or 0)
            profit = rev - cost
            margin = (profit / rev * 100) if rev else 0
            data.append([i, row['medicine__medicine_name'], row['medicine__strength'],
                          row['units'], f'₱{rev:,.2f}', f'₱{cost:,.2f}',
                          f'₱{profit:,.2f}', f'{margin:.1f}%'])
            color = GREEN if profit >= 0 else RED
            ts.add('TEXTCOLOR', (6, i), (6, i), color)
            ts.add('FONTNAME',  (6, i), (6, i), 'Helvetica-Bold')
        col_widths = [1*cm, 6*cm, 2.5*cm, 2*cm, 3*cm, 3*cm, 3.5*cm, 2.5*cm]
    else:
        data       = [['No data']]
        ts         = make_table_style(1, 1)
        col_widths = None

    if len(data) > 1 or data[0][0] != 'No data':
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(ts)
        elements.append(table)
    else:
        elements.append(Paragraph('No data available for this report.', body_style))

    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(
        f'— End of Report — | MHARSMC Electronic Pharmacy Inventory System | {timezone.now().strftime("%B %d, %Y")}',
        sub_style
    ))

    doc.build(elements)
    buffer.seek(0)
    filename = f'MHARSMC_{report_type}_{today}.pdf'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(buffer.read())
    return response


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'consumption_report'), login_url='dashboard')
def consumption_report(request):
    # Helper function to format name with middle initial
    def get_formatted_name(user):
        """Return formatted name: FIRST MIDDLEINITIAL. LAST (all caps)"""
        first_name = user.first_name or ""
        last_name = user.last_name or ""
        
        # Get middle name from profile
        middle_name = ""
        if hasattr(user, 'profile'):
            middle_name = user.profile.middle_name or ""
        
        # Get middle initial
        middle_initial = f"{middle_name[0].upper()}." if middle_name else ""
        
        # Build the formatted name
        if middle_initial:
            formatted = f"{first_name.upper()} {middle_initial} {last_name.upper()}"
        else:
            formatted = f"{first_name.upper()} {last_name.upper()}"
        
        return formatted.strip() or user.username.upper()

    # ── 1. Resolve selected month/year ────────────────────────────────────────
    today = timezone.now().date()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            raise ValueError
    except (ValueError, TypeError):
        year, month = today.year, today.month
 
    # Use timezone-aware datetimes to avoid naive datetime warnings
    period_start = timezone.make_aware(timezone.datetime(year, month, 1))
    if month == 12:
        period_end = timezone.make_aware(timezone.datetime(year + 1, 1, 1))
    else:
        period_end = timezone.make_aware(timezone.datetime(year, month + 1, 1))
 
    # ── 2. All active medicines (ordered by medicine_name) ────────────────────
    medicines = Medicine.objects.filter(is_active=True).select_related('category')
 
    rows = []
    for med in medicines:
 
        # ── Previous Balance ─────────────────────────────────────────────────
        # The quantity_after of the LAST stock movement BEFORE the period start.
        last_movement = (
            StockMovement.objects
            .filter(medicine=med, created_at__lt=period_start)
            .order_by('-created_at')
            .first()
        )
        if last_movement:
            previous_balance = last_movement.quantity_after
        else:
            # No movements before this month — reconstruct from current stock
            # by reversing movements that happened during or after the period.
            movements_since = StockMovement.objects.filter(
                medicine=med, created_at__gte=period_start
            )
            delta = sum(m.quantity for m in movements_since)  # signed integer
            previous_balance = max(med.stock_quantity - delta, 0)
 
        # ── Received during Month ────────────────────────────────────────────
        received_qs = StockMovement.objects.filter(
            medicine=med,
            movement_type__in=['in', 'return'],
            created_at__gte=period_start,
            created_at__lt=period_end,
        ).aggregate(total=Sum('quantity'))
        received = received_qs['total'] or 0
 
        # ── Consumed during Month ────────────────────────────────────────────
        # a) Dispensing — DispensingItem is the source of truth for dispensed qty.
        dispensed_qs = DispensingItem.objects.filter(
            medicine=med,
            dispensing__created_at__gte=period_start,
            dispensing__created_at__lt=period_end,
        ).aggregate(total=Sum('quantity'))
        dispensed = dispensed_qs['total'] or 0
 
        # b) Manual outflows ONLY — expired / damaged removals via Stock Management.
        #    'out' is intentionally excluded: dispensing already creates a
        #    StockMovement(type='out'), so counting it here would double the total.
        manual_out_qs = StockMovement.objects.filter(
            medicine=med,
            movement_type__in=['expired', 'damaged'],
            created_at__gte=period_start,
            created_at__lt=period_end,
        ).aggregate(total=Sum('quantity'))
        manual_out = abs(manual_out_qs['total'] or 0)
 
        consumed = dispensed + manual_out
 
        # ── Balance on Hand ──────────────────────────────────────────────────
        balance_on_hand = previous_balance + received - consumed
 
        # ── Unit of Measure ──────────────────────────────────────────────────
        uom = med.dosage_form  # tablet / capsule / vial / etc.
 
        rows.append({
            'item_code':        med.barcode or f"MED-{med.pk:04d}",
            'description':      f"{med.medicine_name} {med.strength}",
            'unit_of_measure':  uom,
            'previous_balance': previous_balance,
            'received':         received,
            'consumed':         consumed,
            'balance_on_hand':  balance_on_hand,
        })
 
    # ── 3. Compute totals ─────────────────────────────────────────────────────
    total_received = sum(r['received']         for r in rows)
    total_consumed = sum(r['consumed']         for r in rows)
    total_balance  = sum(r['balance_on_hand']  for r in rows)
 
    # ── 4. Build month selector options ──────────────────────────────────────
    month_choices    = [(i, calendar.month_name[i]) for i in range(1, 13)]
    year_choices     = list(range(today.year - 3, today.year + 2))
    last_day_of_month = calendar.monthrange(year, month)[1]
 
    # Use the formatted name function
    submitted_by = get_formatted_name(request.user)
    position = request.user.profile.position if hasattr(request.user, 'profile') else '—'

    context = {
        'rows':              rows,
        'selected_year':     year,
        'selected_month':    month,
        'month_name':        calendar.month_name[month],
        'month_choices':     month_choices,
        'year_choices':      year_choices,
        'period_start':      period_start,
        'period_end':        period_end,
        'today':             today,
        'total_received':    total_received,
        'total_consumed':    total_consumed,
        'total_balance':     total_balance,
        'last_day_of_month': last_day_of_month,
        'submitted_by':      submitted_by,
        'position':          position,
    }
    return render(request, 'consumption_report.html', context)



# ─────────────────────────────────────────────────────────────────────────────
# DISPENSING
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'dispensing'), login_url='dashboard')
def dispensing_view(request):
    today = timezone.now().date()

    today_txns    = Dispensing.objects.filter(created_at__date=today)
    txn_today     = today_txns.count()
    units_today   = DispensingItem.objects.filter(
                        dispensing__created_at__date=today).aggregate(
                        t=Sum('quantity'))['t'] or 0
    revenue_today = today_txns.aggregate(t=Sum('total_amount'))['t'] or 0
    month_start   = today.replace(day=1)
    txn_month     = Dispensing.objects.filter(created_at__date__gte=month_start).count()

    per_page    = 10
    txn_qs      = Dispensing.objects.select_related('pharmacist').prefetch_related('items').all()
    paginator   = Paginator(txn_qs, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    context = {
        'txn_today':     txn_today,
        'units_today':   units_today,
        'revenue_today': revenue_today,
        'txn_month':     txn_month,
        'transactions':  page_obj.object_list,
        'page_obj':      page_obj,
        'paginator':     paginator,
    }
    return render(request, 'dispensing.html', context)


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'dispensing'), login_url='dashboard')
@transaction.atomic
def dispensing_create(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method.'}, status=405)

    try:
        data            = json.loads(request.body)
        items_data      = data.get('items', [])
        customer_name   = data.get('customer_name', '').strip()
        prescription_no = data.get('prescription_no', '').strip()
        notes           = data.get('notes', '').strip()
        discount_amount = float(data.get('discount_amount', 0))
        doctor_pk       = data.get('doctor_pk', '')

        # ── iHOMIS patient link (optional) ────────────────────────────────────
        # Sent from the frontend when the pharmacist selects a patient from iHOMIS.
        # ihomis_hpatkey is the hperson.hpatkey (integer PK in iHOMIS).
        # ihomis_hpercode is the hospital number — cached locally so we don't
        # need a live iHOMIS query just to display it.
        # Safe int conversion — prevents string hospital numbers from slipping in
        ihomis_hpatkey  = str(data.get('ihomis_hpatkey',  '') or '').strip() or None
        ihomis_hpercode = str(data.get('ihomis_hpercode', '') or '').strip()

        # Default customer name to Walk-in if nothing was provided
        if not customer_name:
            customer_name = 'Walk-in'

        # Safely handle doctor_pk
        if doctor_pk is not None:
            doctor_pk = str(doctor_pk).strip()
        else:
            doctor_pk = ''

        if not items_data:
            return JsonResponse({'error': 'Cart is empty.'}, status=400)

        # ── Validate all items before any DB writes ───────────────────────────
        validated   = []
        has_s2_item = False

        for item in items_data:
            pk  = item.get('medicine_pk')
            qty = int(item.get('quantity', 0))
            if qty <= 0:
                return JsonResponse({'error': 'Quantity must be > 0.'}, status=400)
            medicine = get_object_or_404(Medicine, pk=pk, is_active=True)
            if qty > medicine.stock_quantity:
                return JsonResponse({
                    'error': f'Insufficient stock for "{medicine.medicine_name}". '
                             f'Available: {medicine.stock_quantity}.'
                }, status=400)
            if medicine.is_s2_regulated:
                has_s2_item = True
            validated.append({
                'medicine':   medicine,
                'qty':        qty,
                'unit_price': float(item.get('unit_price', medicine.selling_price)),
            })

        # ── S2 hard block ─────────────────────────────────────────────────────
        doctor = None
        if has_s2_item:
            if not doctor_pk:
                return JsonResponse({
                    'error': 'A prescribing doctor with a valid S2 license is required '
                             'for S2-regulated medicines.'
                }, status=400)
            try:
                try:
                    doctor_pk_int = int(doctor_pk)
                    doctor = Doctor.objects.get(pk=doctor_pk_int, is_active=True)
                except (ValueError, TypeError):
                    doctor = Doctor.objects.get(pk=doctor_pk, is_active=True)
            except Doctor.DoesNotExist:
                return JsonResponse({
                    'error': 'Selected doctor not found or is inactive.'
                }, status=400)

            if not doctor.s2_is_valid_for_dispensing:
                status = doctor.s2_license_status
                if status == 'none':
                    reason = f'Dr. {doctor.get_full_name()} has no S2 license on record.'
                elif status == 'expired':
                    reason = (f'Dr. {doctor.get_full_name()}\'s S2 license expired on '
                              f'{doctor.s2_expiry_date}.')
                else:
                    reason = f'Dr. {doctor.get_full_name()}\'s S2 license is not valid.'
                return JsonResponse({'error': reason}, status=400)
        # ─────────────────────────────────────────────────────────────────────

        subtotal = sum(v['unit_price'] * v['qty'] for v in validated)
        discount = min(discount_amount, subtotal)
        total    = subtotal - discount

        # ── Create transaction ────────────────────────────────────────────────
        txn = Dispensing.objects.create(
            customer_name   = customer_name or 'Walk-in',
            prescription_no = prescription_no,
            notes           = notes,
            subtotal_amount = subtotal,
            discount_amount = discount,
            total_amount    = total,
            pharmacist      = request.user,
            doctor          = doctor,
            # iHOMIS link — stored as plain integers/strings, no cross-DB FK
            ihomis_hpatkey  = ihomis_hpatkey or '',
            ihomis_hpercode = ihomis_hpercode or '',
        )

        # ── Create line items & deduct stock (FEFO) ───────────────────────────
        receipt_items = []
        for v in validated:
            medicine = v['medicine']
            qty      = v['qty']
            price    = v['unit_price']

            DispensingItem.objects.create(
                dispensing = txn,
                medicine   = medicine,
                quantity   = qty,
                unit_price = price,
            )

            _deduct_fefo(
                medicine, qty, request.user,
                movement_type   = 'out',
                dispensed_to    = customer_name,
                prescription_no = prescription_no,
                reference_no    = f'TXN-{txn.pk}',
                notes           = (
                    f'Dispensing #{txn.pk}'
                    + (f' — S2 Rx by Dr. {doctor.get_full_name()}' if doctor else '')
                    + (f' — {notes}' if notes else '')
                ),
            )

            receipt_items.append({
                'name':       medicine.medicine_name,
                'strength':   medicine.strength,
                'qty':        qty,
                'unit_price': price,
                'subtotal':   price * qty,
                'is_s2':      medicine.is_s2_regulated,
            })

        # ── Build doctor info for receipt ─────────────────────────────────────
        doctor_info = None
        if doctor:
            doctor_info = {
                'name':           f'Dr. {doctor.get_full_name()}',
                'specialization': doctor.specialization,
                'prc_license_no': doctor.prc_license_no,
                's2_license_no':  doctor.s2_license_no,
            }

        return JsonResponse({
            'success': True,
            'receipt': {
                'pk':              txn.pk,
                'date':            txn.created_at.strftime('%B %d, %Y %I:%M %p'),
                'pharmacist':      request.user.get_full_name() or request.user.username,
                'customer_name':   customer_name,
                'prescription_no': prescription_no,
                'items':           receipt_items,
                'subtotal':        subtotal,
                'discount':        discount,
                'total':           total,
                'doctor':          doctor_info,
                # Echo back the linked patient for frontend confirmation
                'ihomis_hpatkey':  ihomis_hpatkey,
                'ihomis_hpercode': ihomis_hpercode,
            }
        })

    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'dispensing'), login_url='dashboard')
def dispensing_receipt(request, pk):
    txn = get_object_or_404(
        Dispensing.objects.select_related('pharmacist').prefetch_related('items__medicine'),
        pk=pk
    )
    items = [
        {
            'name':       item.medicine.medicine_name,
            'strength':   item.medicine.strength,
            'qty':        item.quantity,
            'unit_price': float(item.unit_price),
            'subtotal':   float(item.subtotal),
        }
        for item in txn.items.all()
    ]

    doctor_info = None
    if txn.doctor:
        d = txn.doctor
        doctor_info = {
            'name':           f'Dr. {d.get_full_name()}',
            'specialization': d.specialization,
            'prc_license_no': d.prc_license_no,
            's2_license_no':  d.s2_license_no,
        }

    return JsonResponse({
        'pk':              txn.pk,
        'date':            txn.created_at.strftime('%B %d, %Y %I:%M %p'),
        'pharmacist':      txn.pharmacist.get_full_name() if txn.pharmacist else '—',
        'customer_name':   txn.customer_name,
        'prescription_no': txn.prescription_no,
        'items':           items,
        'subtotal':        float(txn.subtotal_amount),
        'discount':        float(txn.discount_amount),
        'total':           float(txn.total_amount),
        'doctor':          doctor_info,
    })



# views_dispensing_card.py
# Add these views to your existing views.py (or import from here).
# They power: patient search, dispensing card page, and printable card.
# ─────────────────────────────────────────────────────────────────────────────
# AJAX: Search iHOMIS patients
# URL:  /dispensing/patient-search/?q=<term>
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@require_GET
def patient_search_ajax(request):
    """
    Search iHOMIS hperson by hospital number, last name, or first name.
    Returns up to 15 results as JSON.
    """
    q = request.GET.get('q', '').strip()

    if len(q) < 2:
        return JsonResponse({'results': []})

    try:
        qs = HPerson.objects.using('ihomis_plus').filter(
            Q(hpercode__icontains=q)    |
            Q(hpatcode__icontains=q)    |
            Q(patlast__icontains=q)     |
            Q(patfirst__icontains=q)    |
            Q(patmedno__icontains=q)
        ).order_by('patlast', 'patfirst')[:15]

        results = []
        for p in qs:
            results.append({
                'hpatkey':         str(p.hpatkey),
                'hospital_number': p.hospital_number,
                'full_name':       p.full_name,
                'full_name_formal':p.full_name_formal,
                'birthdate':       p.patbdate.strftime('%B %d, %Y') if p.patbdate else '—',
                'sex':             p.sex_display,
                'address':         p.address,
                'phicnum':         p.phicnum or '—',
                'patmedno':        p.patmedno or '—',
            })
    except Exception as e:
        # iHOMIS DB might be unreachable — graceful fallback
        return JsonResponse({'results': [], 'error': str(e)})

    return JsonResponse({'results': results})


# ─────────────────────────────────────────────────────────────────────────────
# AJAX: Fetch one iHOMIS patient by hpatkey
# URL:  /dispensing/patient/<hpatkey>/info/
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@require_GET
def patient_info_ajax(request, hpatkey):
    try:
        p = HPerson.objects.using('ihomis_plus').get(pk=hpatkey)
        return JsonResponse({
            'hpatkey':         p.hpatkey,
            'hospital_number': p.hospital_number,
            'full_name':       p.full_name,
            'full_name_formal':p.full_name_formal,
            'birthdate':       p.patbdate.strftime('%B %d, %Y') if p.patbdate else '—',
            'birthdate_iso':   p.patbdate.isoformat() if p.patbdate else '',
            'sex':             p.sex_display,
            'address':         p.address,
            'phicnum':         p.phicnum or '—',
            'patmedno':        p.patmedno or '—',
            'civil_status':    p.patcstat or '—',
            'contact':         p.pattelno or '—',
        })
    except HPerson.DoesNotExist:
        return JsonResponse({'error': 'Patient not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: Dispensing Card list (search & select patient)
# URL:  /dispensing/card/
# ─────────────────────────────────────────────────────────────────────────────
@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'dispensing_card'), login_url='dashboard')
def dispensing_card_list(request):
    """
    Landing page for the Dispensing Card module.
    Shows a patient search box. On selection, loads the card via AJAX or
    redirects to the card detail page.
    """
    return render(request, 'dispensing_card.html', {
        'page': 'dispensing_card',
    })


# ─────────────────────────────────────────────────────────────────────────────
# PAGE/AJAX: Dispensing Card for one patient
# URL:  /dispensing/card/<hpatkey>/
# ─────────────────────────────────────────────────────────────────────────────
@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'dispensing_card'), login_url='dashboard')
def dispensing_card_detail(request, hpatkey):
    hpatkey_str = str(hpatkey).strip()
 
    # Fetch patient from iHOMIS
    try:
        patient = HPerson.objects.using('ihomis_plus').get(pk=hpatkey_str)
    except HPerson.DoesNotExist:
        print(f"[DEBUG] Patient not found: {hpatkey_str}")
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'error': f'Patient not found: {hpatkey_str}'}, status=404)
        return render(request, 'dispensing_card.html', {
            'error': f'Patient not found in iHOMIS.',
            'page':  'dispensing_card',
        })
    except Exception as e:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'error': f'iHOMIS unavailable: {e}'}, status=503)
        return render(request, 'dispensing_card.html', {
            'error': f'Cannot reach iHOMIS database: {e}',
            'page':  'dispensing_card',
        })
 
    # Fetch dispensing records — try both the hpatkey and the hospital number
    # in case some records were saved with hpercode instead of hpatkey
    transactions_qs = (
        Dispensing.objects
        .filter(ihomis_hpatkey=hpatkey_str)
        .prefetch_related('items__medicine', 'doctor', 'pharmacist')
        .order_by('-created_at')
    )
 
    # Fallback: also check by hospital number (hpercode) in case that was stored
    if not transactions_qs.exists():
        hpercode = patient.hospital_number  # uses the property
        print(f"[DEBUG] No transactions found with hpatkey, trying hpercode {repr(hpercode)}")
        transactions_qs = (
            Dispensing.objects
            .filter(ihomis_hpercode=hpercode)
            .prefetch_related('items__medicine', 'doctor', 'pharmacist')
            .order_by('-created_at')
        )
 
    stats = transactions_qs.aggregate(
        total_transactions=Count('id'),
        total_amount=Sum('total_amount'),
        total_units=Sum('items__quantity'),
    )
 
    txn_list = []
    for txn in transactions_qs:
        items = []
        for item in txn.items.all():
            items.append({
                'medicine_name': item.medicine.medicine_name,
                'generic_name':  item.medicine.generic_name,
                'strength':      item.medicine.strength,
                'dosage_form':   item.medicine.get_dosage_form_display(),
                'quantity':      item.quantity,
                'unit_price':    float(item.unit_price),
                'subtotal':      float(item.subtotal),
                'is_s2':         item.medicine.is_s2_regulated,
            })
        txn_list.append({
            'pk':              txn.pk,
            'date':            txn.created_at.strftime('%B %d, %Y'),
            'time':            txn.created_at.strftime('%I:%M %p'),
            'datetime_iso':    txn.created_at.isoformat(),
            'prescription_no': txn.prescription_no or '—',
            'notes':           txn.notes or '',
            'subtotal':        float(txn.subtotal_amount),
            'discount':        float(txn.discount_amount),
            'total':           float(txn.total_amount),
            'pharmacist':      (txn.pharmacist.get_full_name() or txn.pharmacist.username)
                               if txn.pharmacist else '—',
            'doctor':          f'Dr. {txn.doctor.get_full_name()}' if txn.doctor else '—',
            'doctor_s2':       txn.doctor.s2_license_no if txn.doctor else '',
            'items':           items,
        })
 
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'patient': {
                'hpatkey':          patient.hpatkey,
                'hospital_number':  patient.hospital_number,
                'full_name':        patient.full_name,
                'full_name_formal': patient.full_name_formal,
                'birthdate':        patient.patbdate.strftime('%B %d, %Y') if patient.patbdate else '—',
                'sex':              patient.sex_display,
                'address':          patient.address,
                'phicnum':          patient.phicnum or '—',
                'patmedno':         patient.patmedno or '—',
            },
            'stats': {
                'total_transactions': stats['total_transactions'] or 0,
                'total_amount':       float(stats['total_amount'] or 0),
                'total_units':        stats['total_units'] or 0,
            },
            'transactions': txn_list,
        })
 
    return render(request, 'dispensing_card.html', {
        'page':         'dispensing_card',
        'patient':      patient,
        'transactions': txn_list,
        'stats':        stats,
        'hpatkey':      hpatkey_str,
    })

# ─────────────────────────────────────────────────────────────────────────────
# AJAX: Link an iHOMIS patient to a dispensing transaction
# URL:  /dispensing/<pk>/link-patient/   POST
# Called from the dispensing page when pharmacist selects a patient
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def link_patient_to_dispensing(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    txn = get_object_or_404(Dispensing, pk=pk)
    try:
        body     = json.loads(request.body)
        hpatkey  = body.get('hpatkey')
        hpercode = body.get('hpercode', '')

        if not hpatkey:
            return JsonResponse({'error': 'hpatkey required'}, status=400)

        txn.ihomis_hpatkey  = hpatkey
        txn.ihomis_hpercode = hpercode
        # Also update customer_name from iHOMIS for legacy display
        if body.get('full_name'):
            txn.customer_name = body['full_name']
        txn.save(update_fields=['ihomis_hpatkey', 'ihomis_hpercode', 'customer_name'])

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)




# ─────────────────────────────────────────────────────────────────────────────
# CATEGORY — AJAX search + create
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def category_search_ajax(request):
    q  = request.GET.get('q', '').strip()
    qs = MedicineCategory.objects.all()
    if q:
        qs = qs.filter(name__icontains=q)
    qs = qs.order_by('name')[:20]
    return JsonResponse({'results': [{'pk': c.pk, 'name': c.name} for c in qs]})


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'inventory'), login_url='dashboard')
def category_create_ajax(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method.'}, status=405)
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
    except (json.JSONDecodeError, AttributeError):
        name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'error': 'Category name cannot be empty.'}, status=400)
    if MedicineCategory.objects.filter(name__iexact=name).exists():
        existing = MedicineCategory.objects.get(name__iexact=name)
        return JsonResponse({'pk': existing.pk, 'name': existing.name, 'existed': True})
    category = MedicineCategory.objects.create(name=name)
    return JsonResponse({'pk': category.pk, 'name': category.name, 'existed': False})


# ─────────────────────────────────────────────────────────────────────────────
# BATCH MANAGEMENT — AJAX edit + soft-delete
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'stock_management'), login_url='dashboard')
def batch_edit_view(request, pk):
    """
    AJAX POST — edit a MedicineStock batch's correctable fields:
    batch_number, expiry_date, purchase_price, supplier.
    Quantity is NOT editable here — use Stock In / Adjust for that.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method.'}, status=405)
    try:
        batch = get_object_or_404(MedicineStock, pk=pk)
        data  = json.loads(request.body)

        batch.batch_number   = data.get('batch_number', batch.batch_number)
        batch.purchase_price = data.get('purchase_price', batch.purchase_price)
        expiry = data.get('expiry_date', '')
        batch.expiry_date = expiry if expiry else None
        supplier_id = data.get('supplier_id', '')
        batch.supplier_id = supplier_id if supplier_id else None
        batch.save()

        return JsonResponse({
            'success':        True,
            'batch_number':   batch.batch_number,
            'expiry_date':    str(batch.expiry_date) if batch.expiry_date else '',
            'purchase_price': str(batch.purchase_price),
            'supplier_id':    batch.supplier_id or '',
            'message':        f'Batch "{batch.batch_number or pk}" updated successfully.',
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required(login_url='login')
@user_passes_test(lambda u: can_access_module(u, 'stock_management'), login_url='dashboard')
def batch_delete_view(request, pk):
    """
    AJAX POST — soft-delete a MedicineStock batch by zeroing its quantity.
    The batch row is kept for audit trail (StockMovement FK references it).
    Creates a StockMovement log entry to record the removal.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method.'}, status=405)
    try:
        batch    = get_object_or_404(MedicineStock, pk=pk)
        medicine = batch.medicine
        reason   = json.loads(request.body).get('reason', '').strip() if request.body else ''

        if batch.quantity == 0:
            return JsonResponse({'error': 'Batch is already empty.'}, status=400)

        old_total = medicine.stock_quantity
        removed   = batch.quantity
        batch.quantity = 0
        batch.save()
        new_total = medicine.stock_quantity   # recomputed via property

        StockMovement.objects.create(
            medicine        = medicine,
            batch           = batch,
            movement_type   = 'adjust',
            quantity        = -removed,
            quantity_before = old_total,
            quantity_after  = new_total,
            batch_number    = batch.batch_number,
            expiry_date     = batch.expiry_date,
            notes           = reason or f'Batch removed via Batch Management (Batch: {batch.batch_number or pk})',
            performed_by    = request.user,
        )

        return JsonResponse({
            'success':        True,
            'removed':        removed,
            'new_total_stock':new_total,
            'message':        f'Batch "{batch.batch_number or pk}" removed. {removed} units deducted.',
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    


# ─────────────────────────────────────────────────────────────────────────────
# ACCESS MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def access_management_view(request):
    """
    Admin-only page. Shows nothing to configure at the role level —
    individual user permissions are managed inside User Management.
    This view exists as a landing page and overview.
    """
    users_with_perms = User.objects.select_related(
        'profile', 'permissions'
    ).filter(is_active=True).order_by('first_name', 'last_name')

    # Ensure every user has a permissions row
    for u in users_with_perms:
        UserPermission.objects.get_or_create(user=u)

    users_with_perms = User.objects.select_related(
        'profile', 'permissions'
    ).filter(is_active=True).order_by('first_name', 'last_name')

    context = {
        'users_with_perms': users_with_perms,
        'module_choices':   UserPermission.MODULE_CHOICES,
    }
    return render(request, 'access_management.html', context)